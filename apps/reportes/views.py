import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from copy import copy

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.conf import settings
from decimal import Decimal, ROUND_HALF_UP

from apps.academico.models import Aula, PeriodoLectivo, Matricula, AsignacionAcademica, Nota, Curso, EvaluacionActitudinal, Evaluacion
from apps.academico.services import calcular_matriz_vigesimal, obtener_consolidado_aula_maestro
from django.db.models import Prefetch
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

@login_required
def consolidado_notas_admin(request):
    aulas = Aula.objects.all().order_by('nivel', 'grado', 'seccion')
    aula_id = request.GET.get('aula_id')
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    bimestre_actual = request.GET.get('bimestre', periodo_actual.bimestre_actual if periodo_actual else 'I')
    origen = request.GET.get('origen', '')
    asignacion_id = request.GET.get('asignacion_id')
    
    asignaciones = []
    aula_seleccionada = None
    total_alumnos = 0
    promedio_aula = 0.0
    top_estudiantes = []
    todos_estudiantes = []
    asignacion_seleccionada = None
    
    # 💥 INICIALIZAMOS LA VARIABLE PARA LA TABLA ACTITUDINAL
    matriculas = [] 
    
    if aula_id:
        aula_seleccionada = get_object_or_404(Aula, id=aula_id)
        asignaciones = AsignacionAcademica.objects.filter(aula=aula_seleccionada, periodo=periodo_actual).select_related('curso', 'personal')
        total_alumnos = Matricula.objects.filter(aula=aula_seleccionada, periodo=periodo_actual).count()
        
        if asignaciones.exists():
            asignacion_seleccionada = asignaciones.filter(id=asignacion_id).first() if asignacion_id else asignaciones.first()

        # ====================================================
        # 💥 NUEVO: PREFETCH PARA LA MATRIZ ACTITUDINAL
        # ====================================================
        actitudinal_prefetch = Prefetch(
            'actitudinales', 
            queryset=EvaluacionActitudinal.objects.filter(bimestre=bimestre_actual),
            to_attr='eval_actitudinal'
        )
        
        matriculas = Matricula.objects.filter(
            aula=aula_seleccionada, periodo=periodo_actual
        ).select_related('estudiante').prefetch_related(actitudinal_prefetch).order_by('estudiante__apellidos', 'estudiante__nombres')
        # ====================================================

        # ----------------------------------------------------
        # LLAMADO AL MOTOR MAESTRO PARA EXTRAER ANALÍTICAS
        # ----------------------------------------------------
        data_alumnos, _ = obtener_consolidado_aula_maestro(aula_seleccionada, periodo_actual)
        
        # 💥 NUEVO: Vincular el promedio actitudinal maestro redondeado
        for m in matriculas:
            if m.id in data_alumnos:
                # Extraemos el valor ya procesado (ej. 15, 18, etc.)
                m.actitudinal_redondeado = data_alumnos[m.id]['comportamiento'].get(bimestre_actual)
            else:
                m.actitudinal_redondeado = None
        
        top_estudiantes_raw = []
        suma_promedios_aula = 0
        alumnos_validos = 0
        
        for data in data_alumnos.values():
            mat = data['matricula']
            mat.puntaje_total = data['puntajes_bimestre'].get(bimestre_actual, 0)
            top_estudiantes_raw.append(mat)
            
            promedio = data['promedios_bimestre'].get(bimestre_actual, 0)
            if promedio > 0:
                suma_promedios_aula += promedio
                alumnos_validos += 1
                
        # Top 3 (Cuadro de Honor exacto)
        top_estudiantes_raw.sort(key=lambda x: x.puntaje_total, reverse=True)
        top_estudiantes = [est for est in top_estudiantes_raw if est.puntaje_total > 0][:3]
        
        todos_estudiantes = [est for est in top_estudiantes_raw if est.puntaje_total > 0]
        
        # Promedio del Aula exacto
        promedio_aula = (suma_promedios_aula / alumnos_validos) if alumnos_validos > 0 else 0.0

    # Retenemos el llamado a la matriz individual para poblar tu tabla inferior de cursos
    contexto_matriz = calcular_matriz_vigesimal(asignacion_seleccionada, aula_seleccionada, bimestre_actual)
    
    context = {
        'aulas': aulas,
        'asignaciones': asignaciones,
        'aula_seleccionada': aula_seleccionada,
        'bimestre': bimestre_actual,
        'origen': origen,
        'total_alumnos': total_alumnos,
        'promedio_aula': promedio_aula,
        'top_estudiantes': top_estudiantes,
        'todos_estudiantes': todos_estudiantes,
        'asignacion_seleccionada': asignacion_seleccionada,
        'matriculas': matriculas, # 💥 PASAMOS LA VARIABLE AL TEMPLATE
    }
    context.update(contexto_matriz)
    
    return render(request, 'personal/consolidado_notas.html', context)

@login_required
def exportar_matriz_oficial_excel(request, asignacion_id):
    asignacion = get_object_or_404(AsignacionAcademica, id=asignacion_id)
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    bimestre_predeterminado = periodo_actual.bimestre_actual if periodo_actual else 'I'
    bimestre_actual = request.GET.get('bimestre', bimestre_predeterminado)
    en_blanco = request.GET.get('blanco', '0') == '1'

    # 1. Traer datos
    matriculas = Matricula.objects.filter(aula=asignacion.aula, estudiante__estado='Activo').select_related('estudiante').order_by('estudiante__apellidos')
    evaluaciones = asignacion.evaluaciones.filter(bimestre=bimestre_actual)
    notas_db = Nota.objects.filter(evaluacion__in=evaluaciones)
    
    diccionario_notas = {}
    for n in notas_db:
        if n.matricula_id not in diccionario_notas:
            diccionario_notas[n.matricula_id] = {}
        diccionario_notas[n.matricula_id][n.evaluacion_id] = n.valor

    # 💥 CLASIFICACIÓN EXACTA COMO EN LA WEB (Incluye LIBRO y CUADERNO)
    evals_mensual_lc = evaluaciones.filter(tipo__in=['CUADERNO', 'LIBRO'], nombre__icontains='Mensual')
    evals_bimestral_lc = evaluaciones.filter(tipo__in=['CUADERNO', 'LIBRO'], nombre__icontains='Bimestral')
    evals_desafio = evaluaciones.filter(tipo='DESAFIO')
    eval_mensual = evaluaciones.filter(tipo='MENSUAL').first()
    eval_bimestral = evaluaciones.filter(tipo='BIMESTRAL').first()
    evals_simulacro = evaluaciones.filter(tipo='SIMULACRO')

    # 2. Inicializar Excel y Estilos
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Notas {bimestre_actual} Bim"
    
    # Paleta y estilos (Se mantiene intacto tu excelente diseño)
    fill_naranja = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    fill_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_promedios = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid") 
    
    font_titulo = Font(name="Arial", size=14, bold=True)
    font_blanca = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    font_negra_bold = Font(name="Arial", size=9, bold=True)
    font_normal = Font(name="Arial", size=9)
    
    align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_izq = Alignment(horizontal="left", vertical="center")
    align_bottom_center = Alignment(horizontal="center", vertical="bottom")
    align_vertical_bottom = Alignment(horizontal="center", vertical="bottom", textRotation=90)
    
    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    borde_grueso_outer = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # ==========================================
    # 3. CONSTRUCCIÓN DE LA CABECERA PRINCIPAL
    # ==========================================
    ws.row_dimensions[2].height = 35
    ws.merge_cells("A2:W2")
    ws["A2"] = f"INVENTARIO DE GANANCIAS Y PÉRDIDAS DE APRENDIZAJES {asignacion.periodo.anio}"
    ws["A2"].font, ws["A2"].alignment = font_titulo, align_centro
    
    for col in range(1, 24):
        ws.cell(row=2, column=col).border = borde_grueso_outer

    ws.merge_cells("A3:B3")
    ws["A3"] = f"ACTIVIDAD: {asignacion.curso.nombre.upper()}"
    ws.merge_cells("C3:I3")
    ws["C3"] = f"AULA: {asignacion.aula.grado} '{asignacion.aula.seccion}'"
    ws.merge_cells("K3:S3")
    ws["K3"] = f"NIVEL: {asignacion.aula.get_nivel_display().upper()}"
    ws.merge_cells("T3:W3")
    ws["T3"] = f"BIMESTRE: {bimestre_actual}"

    for cell in ["A3", "C3", "K3", "T3"]: ws[cell].font = font_negra_bold

    # ==========================================
    # 4. ESTRUCTURA DE TABLA (FILAS 4, 5 y 6)
    # ==========================================
    ws.row_dimensions[6].height = 160

    ws.merge_cells("A4:A6")
    ws["A4"] = "N°"
    ws["A4"].alignment, ws["A4"].font = align_centro, font_negra_bold

    ws.merge_cells("B4:B6")
    ws["B4"] = "APELLIDOS Y NOMBRES"
    ws["B4"].alignment, ws["B4"].font = align_centro, font_negra_bold

    # -- EVALUACIONES MENSUALES --
    ws.merge_cells("C4:E5")
    ws["C4"] = "EVALUACIONES MENSUALES"
    ws["C4"].fill, ws["C4"].font, ws["C4"].alignment = fill_naranja, font_blanca, align_centro
    ws["C6"], ws["D6"], ws["E6"] = "DESARROLLO DE LIBRO", "DESARROLLO DE TAREAS", "PROM - 1"
    for cell, f, a, ft in zip(["C6","D6","E6"], [fill_blanco, fill_blanco, fill_naranja], [align_vertical_bottom]*3, [font_negra_bold, font_negra_bold, font_blanca]):
        ws[cell].fill, ws[cell].alignment, ws[cell].font = f, a, ft

    # -- EVALUACIONES BIMESTRALES --
    ws.merge_cells("F4:H5")
    ws["F4"] = "EVALUACIONES BIMESTRALES"
    ws["F4"].fill, ws["F4"].font, ws["F4"].alignment = fill_naranja, font_blanca, align_centro
    ws["F6"], ws["G6"], ws["H6"] = "DESARROLLO DE LIBRO", "DESARROLLO DE TAREAS", "PROM - 2"
    for cell, f, a, ft in zip(["F6","G6","H6"], [fill_blanco, fill_blanco, fill_naranja], [align_vertical_bottom]*3, [font_negra_bold, font_negra_bold, font_blanca]):
        ws[cell].fill, ws[cell].alignment, ws[cell].font = f, a, ft

    # -- EVALUACIONES DIARIAS --
    ws.merge_cells("I4:Q4")
    ws["I4"] = "EVALUACIONES DIARIAS"
    ws["I4"].fill, ws["I4"].font, ws["I4"].alignment = fill_naranja, font_blanca, align_centro

    ws.merge_cells("I5:Q5")
    ws["I5"] = "DESAFIO EMPRENDEDOR"
    ws["I5"].fill, ws["I5"].font, ws["I5"].alignment = fill_blanco, font_negra_bold, align_centro

    for i in range(1, 9):
        col_letra = openpyxl.utils.get_column_letter(8 + i)
        ws[f"{col_letra}6"] = str(i)
        ws[f"{col_letra}6"].fill, ws[f"{col_letra}6"].alignment, ws[f"{col_letra}6"].font = fill_blanco, align_bottom_center, font_negra_bold

    ws["Q6"] = "PROM - 3"
    ws["Q6"].fill, ws["Q6"].alignment, ws["Q6"].font = fill_naranja, align_vertical_bottom, font_blanca

    # -- EXÁMENES IMPORTANTES --
    ws.merge_cells("R4:R6")
    ws["R4"] = "CONTROL DE CALIDAD"
    ws["R4"].fill, ws["R4"].alignment, ws["R4"].font = fill_naranja, align_vertical_bottom, font_blanca

    ws.merge_cells("S4:S6")
    ws["S4"] = "ISO INGENIERÍA"
    ws["S4"].fill, ws["S4"].alignment, ws["S4"].font = fill_naranja, align_vertical_bottom, font_blanca

    # -- CONCURSO DE APTITUD --
    ws.merge_cells("T4:V5")
    ws["T4"] = "CONCURSO DE APTITUD"
    ws["T4"].fill, ws["T4"].font, ws["T4"].alignment = fill_naranja, font_blanca, align_centro

    ws["T6"], ws["U6"], ws["V6"] = "CONCURSO DE APTITUD MENSUAL", "CONCURSO DE APTITUD MENSUAL", "PROM - CONCURSO DE APTITUD"
    for cell, f in zip(["T6","U6","V6"], [fill_blanco, fill_blanco, fill_naranja]):
        ws[cell].fill, ws[cell].alignment, ws[cell].font = f, align_vertical_bottom, (font_blanca if f == fill_naranja else font_negra_bold)

    # -- CERTIFICACIÓN DE CALIDAD --
    ws.merge_cells("W4:W6")
    ws["W4"] = "CERTIFICACIÓN DE CALIDAD"
    ws["W4"].fill, ws["W4"].alignment, ws["W4"].font = fill_naranja, align_vertical_bottom, font_blanca

    for row in ws.iter_rows(min_row=4, max_row=6, min_col=1, max_col=23):
        for cell in row: cell.border = borde_fino

    # ==========================================
    # 5. ESCRIBIR DATOS Y MATEMÁTICA ESTRICTA (Desde Fila 7)
    # ==========================================
    fila_actual = 7
    for idx, mat in enumerate(matriculas, 1):
        ws[f"A{fila_actual}"] = idx
        ws[f"B{fila_actual}"] = f"{mat.estudiante.apellidos}, {mat.estudiante.nombres}"
        ws[f"A{fila_actual}"].alignment = align_centro
        ws[f"B{fila_actual}"].alignment = align_izq

        for col_idx in range(1, 24):
            celda = ws.cell(row=fila_actual, column=col_idx)
            celda.border = borde_fino
            
            # 💥 LA MAGIA: Las columnas E(5), H(8), Q(17), R(18), S(19), V(22), W(23)
            if col_idx in [5, 8, 17, 18, 19, 22, 23]:
                celda.fill = fill_promedios
                celda.font = font_negra_bold  # Aplicamos la negrita aquí
            else:
                celda.font = font_normal      # Fuente normal para el resto

            if col_idx > 2: 
                celda.alignment = align_centro

        if not en_blanco:
            notas_alumno = diccionario_notas.get(mat.id, {})
            
            # 💥 Función robusta para obtener una nota individual redondeada
            def obtener_nota(eval_obj):
                if eval_obj and eval_obj.id in notas_alumno and notas_alumno[eval_obj.id] is not None:
                    return int(Decimal(str(notas_alumno[eval_obj.id])).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
                return ""

            # 💥 Función robusta de promedio
            def calcular_promedio(grupo):
                valores = [obtener_nota(e) for e in grupo if obtener_nota(e) != ""]
                if valores:
                    promedio = sum(valores) / len(valores)
                    return int(Decimal(str(promedio)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
                return ""

            # Identificamos evaluaciones específicas para las celdas C, D, F, G
            libro_mensual = evals_mensual_lc.filter(tipo='LIBRO').first()
            cuad_mensual = evals_mensual_lc.filter(tipo='CUADERNO').first()
            libro_bim = evals_bimestral_lc.filter(tipo='LIBRO').first()
            cuad_bim = evals_bimestral_lc.filter(tipo='CUADERNO').first()

            # Calculamos los 6 componentes principales
            prom_mensual_lc = calcular_promedio(evals_mensual_lc)
            prom_bimestral_lc = calcular_promedio(evals_bimestral_lc)
            prom_desafio = calcular_promedio(evals_desafio)
            nota_mensual = obtener_nota(eval_mensual)
            nota_bimestral = obtener_nota(eval_bimestral)
            
            sims = list(evals_simulacro)
            nota_sim1 = obtener_nota(sims[0]) if len(sims) > 0 else ""
            nota_sim2 = obtener_nota(sims[1]) if len(sims) > 1 else ""
            prom_simulacro = calcular_promedio(evals_simulacro)

            # Promedio Final
            componentes = [prom_mensual_lc, prom_bimestral_lc, prom_desafio, nota_mensual, nota_bimestral, prom_simulacro]
            sumatoria_validos = [p for p in componentes if p != ""]
            prom_general = int(Decimal(str(sum(sumatoria_validos) / len(sumatoria_validos))).quantize(Decimal('1'), rounding=ROUND_HALF_UP)) if sumatoria_validos else ""

            # 💥 ESCRITURA EN EXCEL
            ws[f"C{fila_actual}"] = obtener_nota(libro_mensual)
            ws[f"D{fila_actual}"] = obtener_nota(cuad_mensual)
            ws[f"E{fila_actual}"] = prom_mensual_lc
            
            ws[f"F{fila_actual}"] = obtener_nota(libro_bim)
            ws[f"G{fila_actual}"] = obtener_nota(cuad_bim)
            ws[f"H{fila_actual}"] = prom_bimestral_lc

            for i, ev in enumerate(list(evals_desafio)[:8]):
                col_letra = openpyxl.utils.get_column_letter(9 + i)
                ws[f"{col_letra}{fila_actual}"] = obtener_nota(ev)
            
            ws[f"Q{fila_actual}"] = prom_desafio
            ws[f"R{fila_actual}"] = nota_mensual
            ws[f"S{fila_actual}"] = nota_bimestral
            ws[f"T{fila_actual}"] = nota_sim1
            ws[f"U{fila_actual}"] = nota_sim2
            ws[f"V{fila_actual}"] = prom_simulacro
            ws[f"W{fila_actual}"] = prom_general

        fila_actual += 1

    # 6. Ajustar Anchos
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    for col_idx in range(3, 24):
        col_letra = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letra].width = 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Matriz_{asignacion.curso.nombre}_{bimestre_actual}B.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_sabanas_aula_excel(request, aula_id):
    aula = get_object_or_404(Aula, id=aula_id)
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    bimestre_actual = request.GET.get('bimestre', periodo_actual.bimestre_actual if periodo_actual else 'I')
    en_blanco = request.GET.get('blanco', '0') == '1'

    asignaciones = AsignacionAcademica.objects.filter(aula=aula, periodo=periodo_actual).select_related('curso', 'personal')
    matriculas = Matricula.objects.filter(aula=aula, estudiante__estado='Activo').select_related('estudiante').order_by('estudiante__apellidos')

    # 1. Extracción Masiva en Memoria (Evita consultas N+1)
    mapa_notas = {}
    evaluaciones_asig = {}
    
    if not en_blanco:
        evaluaciones = Evaluacion.objects.filter(asignacion__in=asignaciones, bimestre=bimestre_actual)
        for ev in evaluaciones:
            evaluaciones_asig.setdefault(ev.asignacion_id, []).append(ev)
            
        notas_db = Nota.objects.filter(evaluacion__in=evaluaciones, matricula__in=matriculas)
        for n in notas_db:
            if n.matricula_id not in mapa_notas: mapa_notas[n.matricula_id] = {}
            mapa_notas[n.matricula_id][n.evaluacion_id] = n.valor

    # 2. Inicializar Excel y Estilos
    wb = openpyxl.Workbook()
    hoja_default = wb.active 

    fill_naranja = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    fill_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_promedios = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid") 
    
    font_titulo = Font(name="Arial", size=14, bold=True)
    font_blanca = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    font_negra_bold = Font(name="Arial", size=9, bold=True)
    font_normal = Font(name="Arial", size=9)
    
    align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_izq = Alignment(horizontal="left", vertical="center")
    align_bottom_center = Alignment(horizontal="center", vertical="bottom")
    align_vertical_bottom = Alignment(horizontal="center", vertical="bottom", textRotation=90)
    
    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    borde_grueso_outer = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # 3. Iteración por cada curso para crear una pestaña (sheet) nueva
    for asig in asignaciones:
        # El nombre de la pestaña soporta max 31 caracteres en Excel y no permite ciertos símbolos
        nombre_limpio = asig.curso.nombre.replace("/", "-").replace("\\", "-")[:31]
        ws = wb.create_sheet(title=nombre_limpio)
        
        # --- CABECERA PRINCIPAL ---
        ws.row_dimensions[2].height = 35
        ws.merge_cells("A2:W2")
        ws["A2"] = f"INVENTARIO DE GANANCIAS Y PÉRDIDAS DE APRENDIZAJES {periodo_actual.anio}"
        ws["A2"].font, ws["A2"].alignment = font_titulo, align_centro
        
        for col in range(1, 24):
            ws.cell(row=2, column=col).border = borde_grueso_outer

        ws.merge_cells("A3:B3")
        ws["A3"] = f"ACTIVIDAD: {asig.curso.nombre.upper()}"
        ws.merge_cells("C3:I3")
        ws["C3"] = f"AULA: {aula.grado} '{aula.seccion}'"
        ws.merge_cells("K3:S3")
        ws["K3"] = f"NIVEL: {aula.get_nivel_display().upper()}"
        ws.merge_cells("T3:W3")
        ws["T3"] = f"BIMESTRE: {bimestre_actual}"

        for cell in ["A3", "C3", "K3", "T3"]: ws[cell].font = font_negra_bold

        # --- ESTRUCTURA DE TABLA (FILAS 4, 5 y 6) ---
        ws.row_dimensions[6].height = 160

        ws.merge_cells("A4:A6")
        ws["A4"] = "N°"
        ws["A4"].alignment, ws["A4"].font = align_centro, font_negra_bold

        ws.merge_cells("B4:B6")
        ws["B4"] = "APELLIDOS Y NOMBRES"
        ws["B4"].alignment, ws["B4"].font = align_centro, font_negra_bold

        ws.merge_cells("C4:E5")
        ws["C4"] = "EVALUACIONES MENSUALES"
        ws["C4"].fill, ws["C4"].font, ws["C4"].alignment = fill_naranja, font_blanca, align_centro
        ws["C6"], ws["D6"], ws["E6"] = "DESARROLLO DE LIBRO", "DESARROLLO DE TAREAS", "PROM - 1"
        for cell, f, a, ft in zip(["C6","D6","E6"], [fill_blanco, fill_blanco, fill_naranja], [align_vertical_bottom]*3, [font_negra_bold, font_negra_bold, font_blanca]):
            ws[cell].fill, ws[cell].alignment, ws[cell].font = f, a, ft

        ws.merge_cells("F4:H5")
        ws["F4"] = "EVALUACIONES BIMESTRALES"
        ws["F4"].fill, ws["F4"].font, ws["F4"].alignment = fill_naranja, font_blanca, align_centro
        ws["F6"], ws["G6"], ws["H6"] = "DESARROLLO DE LIBRO", "DESARROLLO DE TAREAS", "PROM - 2"
        for cell, f, a, ft in zip(["F6","G6","H6"], [fill_blanco, fill_blanco, fill_naranja], [align_vertical_bottom]*3, [font_negra_bold, font_negra_bold, font_blanca]):
            ws[cell].fill, ws[cell].alignment, ws[cell].font = f, a, ft

        ws.merge_cells("I4:Q4")
        ws["I4"] = "EVALUACIONES DIARIAS"
        ws["I4"].fill, ws["I4"].font, ws["I4"].alignment = fill_naranja, font_blanca, align_centro

        ws.merge_cells("I5:Q5")
        ws["I5"] = "DESAFIO EMPRENDEDOR"
        ws["I5"].fill, ws["I5"].font, ws["I5"].alignment = fill_blanco, font_negra_bold, align_centro

        for i in range(1, 9):
            col_letra = openpyxl.utils.get_column_letter(8 + i)
            ws[f"{col_letra}6"] = str(i)
            ws[f"{col_letra}6"].fill, ws[f"{col_letra}6"].alignment, ws[f"{col_letra}6"].font = fill_blanco, align_bottom_center, font_negra_bold

        ws["Q6"] = "PROM - 3"
        ws["Q6"].fill, ws["Q6"].alignment, ws["Q6"].font = fill_naranja, align_vertical_bottom, font_blanca

        ws.merge_cells("R4:R6")
        ws["R4"] = "CONTROL DE CALIDAD"
        ws["R4"].fill, ws["R4"].alignment, ws["R4"].font = fill_naranja, align_vertical_bottom, font_blanca

        ws.merge_cells("S4:S6")
        ws["S4"] = "ISO INGENIERÍA"
        ws["S4"].fill, ws["S4"].alignment, ws["S4"].font = fill_naranja, align_vertical_bottom, font_blanca

        ws.merge_cells("T4:V5")
        ws["T4"] = "CONCURSO DE APTITUD"
        ws["T4"].fill, ws["T4"].font, ws["T4"].alignment = fill_naranja, font_blanca, align_centro

        ws["T6"], ws["U6"], ws["V6"] = "CONCURSO DE APTITUD MENSUAL", "CONCURSO DE APTITUD MENSUAL", "PROM - CONCURSO DE APTITUD"
        for cell, f in zip(["T6","U6","V6"], [fill_blanco, fill_blanco, fill_naranja]):
            ws[cell].fill, ws[cell].alignment, ws[cell].font = f, align_vertical_bottom, (font_blanca if f == fill_naranja else font_negra_bold)

        ws.merge_cells("W4:W6")
        ws["W4"] = "CERTIFICACIÓN DE CALIDAD"
        ws["W4"].fill, ws["W4"].alignment, ws["W4"].font = fill_naranja, align_vertical_bottom, font_blanca

        for row in ws.iter_rows(min_row=4, max_row=6, min_col=1, max_col=23):
            for cell in row: cell.border = borde_fino

        # --- ESCRIBIR DATOS Y MATEMÁTICA ---
        evals = evaluaciones_asig.get(asig.id, [])
        evals_mensual_lc = [e for e in evals if e.tipo in ['CUADERNO', 'LIBRO'] and 'Mensual' in e.nombre]
        evals_bimestral_lc = [e for e in evals if e.tipo in ['CUADERNO', 'LIBRO'] and 'Bimestral' in e.nombre]
        evals_desafio = [e for e in evals if e.tipo == 'DESAFIO']
        eval_mensual = next((e for e in evals if e.tipo == 'MENSUAL'), None)
        eval_bimestral = next((e for e in evals if e.tipo == 'BIMESTRAL'), None)
        evals_simulacro = [e for e in evals if e.tipo == 'SIMULACRO']

        libro_mensual = next((e for e in evals_mensual_lc if e.tipo == 'LIBRO'), None)
        cuad_mensual = next((e for e in evals_mensual_lc if e.tipo == 'CUADERNO'), None)
        libro_bim = next((e for e in evals_bimestral_lc if e.tipo == 'LIBRO'), None)
        cuad_bim = next((e for e in evals_bimestral_lc if e.tipo == 'CUADERNO'), None)

        fila_actual = 7
        for idx, mat in enumerate(matriculas, 1):
            ws[f"A{fila_actual}"] = idx
            ws[f"B{fila_actual}"] = f"{mat.estudiante.apellidos}, {mat.estudiante.nombres}"
            ws[f"A{fila_actual}"].alignment = align_centro
            ws[f"B{fila_actual}"].alignment = align_izq

            # 💥 APLICANDO FORMATO Y NEGRITAS ESPECÍFICAS
            for col_idx in range(1, 24):
                celda = ws.cell(row=fila_actual, column=col_idx)
                celda.border = borde_fino
                
                # Columnas específicas (E,H,Q,R,S,V,W) en negrita y con fondo especial
                if col_idx in [5, 8, 17, 18, 19, 22, 23]:
                    celda.font = font_negra_bold
                    celda.fill = fill_promedios
                else:
                    celda.font = font_normal

                if col_idx > 2: celda.alignment = align_centro

            if not en_blanco:
                notas_alumno = mapa_notas.get(mat.id, {})
                
                def obtener_nota(eval_obj):
                    if eval_obj and eval_obj.id in notas_alumno and notas_alumno[eval_obj.id] is not None:
                        return int(Decimal(str(notas_alumno[eval_obj.id])).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
                    return ""

                def calcular_promedio(grupo):
                    valores = [obtener_nota(e) for e in grupo if obtener_nota(e) != ""]
                    if valores:
                        promedio = sum(valores) / len(valores)
                        return int(Decimal(str(promedio)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
                    return ""

                prom_mensual_lc = calcular_promedio(evals_mensual_lc)
                prom_bimestral_lc = calcular_promedio(evals_bimestral_lc)
                prom_desafio = calcular_promedio(evals_desafio)
                nota_mensual = obtener_nota(eval_mensual)
                nota_bimestral = obtener_nota(eval_bimestral)
                
                nota_sim1 = obtener_nota(evals_simulacro[0]) if len(evals_simulacro) > 0 else ""
                nota_sim2 = obtener_nota(evals_simulacro[1]) if len(evals_simulacro) > 1 else ""
                prom_simulacro = calcular_promedio(evals_simulacro)

                componentes = [prom_mensual_lc, prom_bimestral_lc, prom_desafio, nota_mensual, nota_bimestral, prom_simulacro]
                sumatoria_validos = [p for p in componentes if p != ""]
                prom_general = int(Decimal(str(sum(sumatoria_validos) / len(sumatoria_validos))).quantize(Decimal('1'), rounding=ROUND_HALF_UP)) if sumatoria_validos else ""

                ws[f"C{fila_actual}"] = obtener_nota(libro_mensual)
                ws[f"D{fila_actual}"] = obtener_nota(cuad_mensual)
                ws[f"E{fila_actual}"] = prom_mensual_lc
                
                ws[f"F{fila_actual}"] = obtener_nota(libro_bim)
                ws[f"G{fila_actual}"] = obtener_nota(cuad_bim)
                ws[f"H{fila_actual}"] = prom_bimestral_lc

                for i, ev in enumerate(list(evals_desafio)[:8]):
                    col_letra = openpyxl.utils.get_column_letter(9 + i)
                    ws[f"{col_letra}{fila_actual}"] = obtener_nota(ev)
                
                ws[f"Q{fila_actual}"] = prom_desafio
                ws[f"R{fila_actual}"] = nota_mensual
                ws[f"S{fila_actual}"] = nota_bimestral
                ws[f"T{fila_actual}"] = nota_sim1
                ws[f"U{fila_actual}"] = nota_sim2
                ws[f"V{fila_actual}"] = prom_simulacro
                ws[f"W{fila_actual}"] = prom_general

            fila_actual += 1

        # --- AJUSTE DE ANCHO DE COLUMNAS ---
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        for col_idx in range(3, 24):
            col_letra = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letra].width = 5

    # 4. Limpieza final: Eliminar la hoja vacía por defecto
    if len(wb.sheetnames) > 1:
        wb.remove(hoja_default)

    # 5. Generar la respuesta de descarga
    tipo_archivo = "Vacias" if en_blanco else "Llenas"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Sabanas_Masivas_{aula.grado}_{aula.seccion}_{tipo_archivo}.xlsx'
    wb.save(response)
    
    return response

# ==========================================
# FUNCIONES AUXILIARES PARA LIBRETAS
# ==========================================
def redondear(valor):
    return int(float(valor) + 0.5) if valor else 0

def obtener_recomendacion(promedio):
    if promedio == "" or promedio == 0 or promedio is None or promedio == "-": return "-"
    promedio = float(promedio)
    if promedio > 17: return "¡ Excelente, Buen trabajo, puedes llegar aún más lejos !"
    elif promedio > 13: return "¡ Buen trabajo, sigue avanzando !"
    elif promedio > 10: return " ¡ A esforzarse más, tú puedes !"
    else: return " ¡ No te desanimes, vamos, confiamos en ti !"

def clasificar_bimestre_asistencia(fecha, bimestre_actual_sistema):
    """ Filtra y agrupa la asistencia según el calendario escolar de forma histórica """
    mes = fecha.month
    if bimestre_actual_sistema == 'I': return 'I'
    if bimestre_actual_sistema == 'II':
        return 'I' if mes in [3, 4, 5, 6] else 'II'
    elif bimestre_actual_sistema == 'III':
        if mes in [3, 4, 5, 6]: return 'I'
        elif mes in [7, 8]: return 'II'
        else: return 'III'
    else:
        if mes in [3, 4, 5, 6]: return 'I'
        elif mes in [7, 8]: return 'II'
        elif mes in [9, 10]: return 'III'
        else: return 'IV'

@login_required
def exportar_libretas_aula_excel(request, aula_id):
    aula = get_object_or_404(Aula, id=aula_id)
    periodo_actual = PeriodoLectivo.objects.get(activo=True)
    bimestre_sistema = periodo_actual.bimestre_actual 
    
    BIMESTRES_ORDEN = ['I', 'II', 'III', 'IV']
    idx_sistema = BIMESTRES_ORDEN.index(bimestre_sistema)
    columnas_notas = {'I': 'F', 'II': 'H', 'III': 'J', 'IV': 'L'}

    # 💥 MAGIA: El Motor Maestro nos entrega la data ya procesada y la configuración exacta.
    data_alumnos, config = obtener_consolidado_aula_maestro(aula, periodo_actual)

    ruta_plantilla = os.path.join(settings.BASE_DIR, 'templates_archivos', config['ruta_plantilla'])
    ruta_logo = os.path.join(settings.BASE_DIR, 'static', 'assets', 'img', 'logo_colegio_libreta.jpg')
    ruta_logo_titulo = os.path.join(settings.BASE_DIR, 'static', 'assets', 'img', 'titulo_libreta_2026.png')
    wb = openpyxl.load_workbook(ruta_plantilla)
    hoja_molde = wb.active 

    for idx, (mat_id, alumno) in enumerate(data_alumnos.items(), 1):
        mat = alumno['matricula']
        ws = wb.copy_worksheet(hoja_molde)
        ws.title = f"{idx:02d}.- {mat.estudiante.apellidos[:18]}"
        ws.views.sheetView[0].showGridLines = False
        
        # Restaurar Formato Condicional
        for cf_range, cf_rules in hoja_molde.conditional_formatting._cf_rules.items():
            for rule in cf_rules:
                ws.conditional_formatting.add(cf_range, copy(rule))

        ws["C4"] = f"{mat.estudiante.apellidos}, {mat.estudiante.nombres}"
        ws["D6"] = mat.estudiante.dni
        ws["J6"] = f"{aula.grado} '{aula.seccion}'"
        ws["M6"] = aula.get_nivel_display()
        
        # Inyección: Áreas
        for area_code, data_area in alumno['areas'].items():
            fila_excel = config['filas'].get(f"ÁREA_{data_area['nombre_display']}") or config['filas'].get(f"ÁREA_{area_code.upper()}") or config['filas'].get(data_area['nombre_display'])
            if fila_excel:
                for idx_b, (bim, col) in enumerate(columnas_notas.items()):
                    nota = data_area['finales'].get(bim, 0)
                    ws[f"{col}{fila_excel}"] = nota if (idx_b <= idx_sistema and nota > 0) else "-"

        # Inyección: Sub-Cursos
        for cursos_list in alumno['cursos_por_area'].values():
            for curso_data in cursos_list:
                fila_excel = config['filas'].get(curso_data['nombre'].upper())
                if fila_excel:
                    for idx_b, (bim, col) in enumerate(columnas_notas.items()):
                        nota = curso_data['notas'].get(bim, 0)
                        ws[f"{col}{fila_excel}"] = nota if (idx_b <= idx_sistema and nota > 0) else "-"

        # Inyección: Talleres
        for taller_nombre, notas_bims in alumno['talleres_cursos'].items():
            fila_excel = config['filas'].get(taller_nombre.upper())
            if fila_excel:
                for idx_b, (bim, col) in enumerate(columnas_notas.items()):
                    nota = notas_bims.get(bim, 0)
                    ws[f"{col}{fila_excel}"] = nota if (idx_b <= idx_sistema and nota > 0) else "-"

        # Inyección: Comportamiento y Asistencia
        fila_comp = config['filas']['ACTITUDINAL']
        for idx_b, (bim, col) in enumerate(columnas_notas.items()):
            if idx_b <= idx_sistema:
                nota_act = alumno['comportamiento'].get(bim)
                ws[f"{col}{fila_comp}"] = nota_act if nota_act else "-"
                
                asis = alumno['asistencias'][bim]
                total_efectivas = asis['P'] + asis['T'] + asis['J']
                ws[f"{col}{config['filas']['ASISTENCIA_TOTAL']}"] = total_efectivas
                ws[f"{col}{config['filas']['ASISTENCIA_J']}"] = asis['J']
                ws[f"{col}{config['filas']['ASISTENCIA_F']}"] = asis['F']
                ws[f"{col}{config['filas']['ASISTENCIA_T']}"] = asis['T']
            else:
                ws[f"{col}{fila_comp}"] = "-"
                ws[f"{col}{config['filas']['ASISTENCIA_TOTAL']}"] = "-"
                ws[f"{col}{config['filas']['ASISTENCIA_J']}"] = "-"
                ws[f"{col}{config['filas']['ASISTENCIA_F']}"] = "-"
                ws[f"{col}{config['filas']['ASISTENCIA_T']}"] = "-"

        # Inyección: Puntaje, Promedio y Orden de Mérito
        for idx_b, (bim, col) in enumerate(columnas_notas.items()):
            if idx_b <= idx_sistema and alumno['count_elements_bimestre'][bim] > 0:
                ws[f"{col}{config['filas']['PUNTAJE']}"] = alumno['puntajes_bimestre'][bim]
                ws[f"{col}{config['filas']['PROMEDIO']}"] = alumno['promedios_bimestre'][bim]
                ws[f"{col}{config['filas']['ORDEN_MERITO']}"] = f"{alumno['orden_merito'][bim]}º"
            else:
                ws[f"{col}{config['filas']['PUNTAJE']}"] = "-"
                ws[f"{col}{config['filas']['PROMEDIO']}"] = "-"
                ws[f"{col}{config['filas']['ORDEN_MERITO']}"] = "-"

        # Inyección: Observaciones dinámicas
        ws[f"D{config['filas']['OBS_I']}"] = obtener_recomendacion(alumno['promedios_bimestre']['I']) if idx_sistema >= 0 and alumno['count_elements_bimestre']['I'] > 0 else "-"
        ws[f"D{config['filas']['OBS_II']}"] = obtener_recomendacion(alumno['promedios_bimestre']['II']) if idx_sistema >= 1 and alumno['count_elements_bimestre']['II'] > 0 else "-"
        ws[f"D{config['filas']['OBS_III']}"] = obtener_recomendacion(alumno['promedios_bimestre']['III']) if idx_sistema >= 2 and alumno['count_elements_bimestre']['III'] > 0 else "-"
        ws[f"D{config['filas']['OBS_IV']}"] = obtener_recomendacion(alumno['promedios_bimestre']['IV']) if idx_sistema >= 3 and alumno['count_elements_bimestre']['IV'] > 0 else "-"

        # 💥 INYECCIÓN DINÁMICA DEL LOGO SOBRE CADA HOJA CLONADA
        if os.path.exists(ruta_logo):
            img_logo = OpenpyxlImage(ruta_logo)
            img_logo.width = 194   # Ancho optimizado en píxeles para tu celda C2
            img_logo.height = 138  # Alto optimizado en píxeles para tu celda C2
            ws.add_image(img_logo, 'C2')
        
        if os.path.exists(ruta_logo_titulo):
            img_logo_titulo = OpenpyxlImage(ruta_logo_titulo)
            img_logo_titulo.width = 630   # Ancho optimizado en píxeles
            img_logo_titulo.height = 74   # Alto optimizado en píxeles

            # --- INICIO DE CONFIGURACIÓN MILIMÉTRICA ---
            
            # 1. ¿Cuántos píxeles quieres mover la imagen respecto a la esquina de la celda D2?
            desplazar_derecha = 190  # Aumenta para mover a la derecha, usa negativos para la izquierda
            desplazar_abajo = 30    # Aumenta para mover hacia abajo, usa negativos para arriba
            
            # 2. Excel usa índices base 0. La celda 'D2' es: 
            #    Columna D = 3 (A=0, B=1, C=2, D=3)
            #    Fila 2 = 1 (Fila 1=0, Fila 2=1)
            marcador = AnchorMarker(
                col=3, 
                colOff=pixels_to_EMU(desplazar_derecha), 
                row=1, 
                rowOff=pixels_to_EMU(desplazar_abajo)
            )
            
            # 3. Convertimos el ancho y alto de la imagen a EMUs
            tamano = XDRPositiveSize2D(
                pixels_to_EMU(img_logo_titulo.width), 
                pixels_to_EMU(img_logo_titulo.height)
            )
            
            # 4. Ensamblamos el ancla y se la inyectamos a la imagen
            img_logo_titulo.anchor = OneCellAnchor(_from=marcador, ext=tamano)
            
            # 5. La agregamos a la hoja de trabajo 
            # (¡OJO! Ya no le pasamos 'D2' aquí porque el ancla ya tiene las coordenadas exactas)
            ws.add_image(img_logo_titulo)

    if len(wb.sheetnames) > 1:
        wb.remove(hoja_molde)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Libretas_{aula.nivel}_{aula.grado}_{aula.seccion}.xlsx'
    wb.save(response)
    return response