import datetime
import json
import requests
import threading
import os

from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.utils.timezone import localtime, now
from httpx import request
from pydantic import ValidationError
from apps.comunicaciones.models import Notificacion, User
from apps.personal.models import Personal
from .models import AsignacionAcademica, Aula, Curso, EntregaSimulacro, Estudiante, EvaluacionActitudinal, EventoCronograma, EvidenciaDocente, HorarioClase, HorarioRecreo, Institucion, MaterialInstitucional, Matricula, PeriodoLectivo, PreguntaSimulacro, SolicitudImpresion, CatalogoMaterial, InventarioAula
from .forms import AsignacionAcademicaForm, AulaForm, CursoForm, EstudianteForm, EventoCronogramaForm, EvidenciaDocenteForm, HorarioClaseForm, InstitucionForm, MaterialInstitucionalForm, PeriodoLectivoForm, PreguntaSimulacroForm, SimulacroForm # Importamos el formulario que acabamos de crear
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_GET, require_POST
from apps.personal.views import obtener_personal_logueado # Nuestra función utilitaria
from .servicios_ia import analizar_rendimiento_estudiante, generar_4_recomendaciones_ia, generar_diagnostico_cualitativo
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
from django.contrib import messages
from .servicios_ml import agrupar_estudiantes_kmeans # Ajusta tu importación
from django.views.decorators.csrf import csrf_exempt
from docxtpl import DocxTemplate, InlineImage, RichText
from docx.shared import Mm  # Para medir el tamaño de la imagen en Milímetros
from .models import Simulacro, Aula # Asegúrate de importar Simulacro y Aula
from django.conf import settings
from apps.comunicaciones.servicios_whatsapp import WhatsAppService

@require_POST
def pausar_alertas_ajax(request):
    """ API para apagar/encender las notificaciones del robot de WhatsApp """
    # Buscamos el periodo lectivo activo
    periodo = PeriodoLectivo.objects.filter(activo=True).first()
    
    if periodo:
        # Capturamos si el switch está en 'true' o 'false'
        estado = request.POST.get('pausar') == 'true'
        periodo.pausar_notificaciones = estado
        periodo.save()
        return JsonResponse({'success': True})
        
    return JsonResponse({'success': False})

def lista_estudiantes(request):
    # Traemos todos los estudiantes de la base de datos
    estudiantes = Estudiante.objects.all()
    
    # Los enviamos al template (HTML)
    context = {
        'estudiantes': estudiantes
    }
    return render(request, 'academico/lista_estudiantes.html', context)

def crear_estudiante(request):
    if request.method == 'POST':
        form = EstudianteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('academico:lista_estudiantes')
    else:
        form = EstudianteForm()
    
    context = {
        'form': form,
        'titulo': 'Nuevo Estudiante',
        'color_clase': 'bg-gradient-primary' # <--- GRIS para crear
    }
    return render(request, 'academico/form_estudiante.html', context)

def editar_estudiante(request, pk):
    estudiante = get_object_or_404(Estudiante, pk=pk)
    if request.method == 'POST':
        form = EstudianteForm(request.POST, instance=estudiante)
        if form.is_valid():
            form.save()
            return redirect('academico:lista_estudiantes')
    else:
        form = EstudianteForm(instance=estudiante)
    
    context = {
        'form': form,
        'titulo': 'Editar Estudiante',
        'color_clase': 'bg-gradient-info' # <--- NARANJA para editar
    }
    return render(request, 'academico/form_estudiante.html', context)

@require_POST
def eliminar_estudiante_ajax(request, pk):
    estudiante = get_object_or_404(Estudiante, pk=pk)
    estudiante.delete()
    return JsonResponse({'status': 'ok', 'message': 'Estudiante eliminado correctamente.'})

@require_POST
def cambiar_estado_estudiante_ajax(request, pk):
    estudiante = get_object_or_404(Estudiante, pk=pk)
    nuevo_estado = request.POST.get('nuevo_estado')
    if nuevo_estado in dict(Estudiante.ESTADOS):
        estudiante.estado = nuevo_estado
        estudiante.save()
        return JsonResponse({'status': 'ok', 'message': 'Estado actualizado correctamente.'})
    return JsonResponse({'status': 'error', 'message': 'Estado no válido.'})

def lista_aulas(request):
    aulas = Aula.objects.all()
    form = AulaForm()
    # Filtramos para que solo traiga a los docentes activos y que sean FIJOS
    docentes = Personal.objects.filter(cargo='DOC', estado='Activo', tipo_contrato='Fijo')
    return render(request, 'academico/lista_aulas.html', {'aulas': aulas, 'form': form, 'docentes': docentes})

def crear_aula(request):
    if request.method == 'POST':
        form = AulaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('academico:lista_aulas')
    else:
        form = AulaForm()
    return render(request, 'academico/form_aula.html', {'form': form, 'titulo': 'Nueva Aula'})

def guardar_aula_ajax(request):
    data = {}
    if request.method == 'POST':
        pk = request.POST.get('aula_id')
        if pk: # Si hay ID, es EDITAR
            aula = Aula.objects.get(pk=pk)
            form = AulaForm(request.POST, instance=aula)
            data['message'] = "¡Aula actualizada correctamente!"
        else: # Si no hay ID, es CREAR
            form = AulaForm(request.POST)
            data['message'] = "¡Aula guardada correctamente!"

        if form.is_valid():
            form.save()
            data['status'] = 'ok'
        else:
            data['status'] = 'error'
            data['errors'] = form.errors
    return JsonResponse(data)

def obtener_aula_data(request, pk):
    aula = Aula.objects.get(pk=pk)
    data = {
        'id': aula.id,
        'nivel': aula.nivel,
        'grado': aula.grado,
        'seccion': aula.seccion,
        'denominacion': aula.denominacion,
        # 💥 NUEVO: Enviamos el ID del tutor (si tiene uno)
        'tutor': aula.tutor.id if aula.tutor else '' 
    }
    return JsonResponse(data)

@require_POST
def eliminar_aula_ajax(request, pk):
    aula = get_object_or_404(Aula, pk=pk)
    # Verificamos si el aula tiene alumnos para evitar errores de integridad (opcional pero recomendado)
    if aula.estudiantes.exists():
        return JsonResponse({
            'status': 'error', 
            'message': 'No puedes eliminar esta aula porque tiene estudiantes asignados.'
        })
    
    aula.delete()
    return JsonResponse({'status': 'ok', 'message': 'Aula eliminada correctamente.'})

def lista_periodos(request):
    periodos = PeriodoLectivo.objects.all()
    form = PeriodoLectivoForm()
    return render(request, 'academico/lista_periodos.html', {'periodos': periodos, 'form': form})

def guardar_periodo_ajax(request):
    pk = request.POST.get('periodo_id')
    if pk:
        periodo = get_object_or_404(PeriodoLectivo, pk=pk)
        form = PeriodoLectivoForm(request.POST, instance=periodo)
    else:
        form = PeriodoLectivoForm(request.POST)

    if form.is_valid():
        # Lógica Pro: Si este periodo se marca como activo, desactivamos todos los demás
        nuevo_periodo = form.save(commit=False)
        if nuevo_periodo.activo:
            PeriodoLectivo.objects.all().update(activo=False)
        nuevo_periodo.save()
        return JsonResponse({'status': 'ok', 'message': 'Periodo guardado correctamente.'})
    return JsonResponse({'status': 'error', 'errors': form.errors})

def obtener_periodo_data(request, pk):
    periodo = get_object_or_404(PeriodoLectivo, pk=pk)
    return JsonResponse({
        'id': periodo.id,
        'anio': periodo.anio,
        'activo': periodo.activo,
    })

@require_POST
def eliminar_periodo_ajax(request, pk):
    periodo = get_object_or_404(PeriodoLectivo, pk=pk)
    periodo.delete()
    return JsonResponse({'status': 'ok', 'message': 'Periodo eliminado.'})

def lista_cursos(request):
    cursos = Curso.objects.all()
    form = CursoForm()
    return render(request, 'academico/lista_cursos.html', {'cursos': cursos, 'form': form})

def guardar_curso_ajax(request):
    data = {}
    if request.method == 'POST':
        pk = request.POST.get('curso_id')
        if pk: # Si hay ID, es EDITAR
            curso = Curso.objects.get(pk=pk)
            form = CursoForm(request.POST, instance=curso)
            data['message'] = "¡Curso actualizado correctamente!"
        else: # Si no hay ID, es CREAR
            form = CursoForm(request.POST)
            data['message'] = "¡Curso guardado correctamente!"

        if form.is_valid():
            form.save()
            data['status'] = 'ok'
        else:
            data['status'] = 'error'
            data['errors'] = form.errors
    return JsonResponse(data)

def obtener_curso_data(request, pk):
    curso = get_object_or_404(Curso, pk=pk)
    return JsonResponse({
        'id': curso.id,
        'nombre': curso.nombre,
        'area': curso.area,  # 💥 ASEGÚRATE DE QUE ESTE CAMPO ESTÉ AQUÍ
        'descripcion': curso.descripcion,
        'activo': curso.activo,
    })

@require_POST
def eliminar_curso_ajax(request, pk):
    curso = get_object_or_404(Curso, pk=pk)
    curso.delete()
    return JsonResponse({'status': 'ok', 'message': 'Curso eliminado.'})

def lista_asignaciones(request):
    asignaciones = AsignacionAcademica.objects.select_related('personal', 'curso', 'aula', 'periodo').all()
    form = AsignacionAcademicaForm()
    return render(request, 'academico/lista_asignaciones.html', {
        'asignaciones': asignaciones, 
        'form': form
    })

def guardar_asignacion_ajax(request):
    pk = request.POST.get('asignacion_id')
    if pk:
        asignacion = get_object_or_404(AsignacionAcademica, pk=pk)
        form = AsignacionAcademicaForm(request.POST, instance=asignacion)
    else:
        form = AsignacionAcademicaForm(request.POST)

    if form.is_valid():
        form.save()
        return JsonResponse({'status': 'ok', 'message': 'Asignación guardada con éxito.'})
    return JsonResponse({'status': 'error', 'errors': form.errors})

def obtener_asignacion_data(request, pk):
    asignacion = get_object_or_404(AsignacionAcademica, pk=pk)
    return JsonResponse({
        'id': asignacion.id,
        'personal': asignacion.personal.id,
        'curso': asignacion.curso.id,
        'aula': asignacion.aula.id,
        'periodo': asignacion.periodo.id,
    })

@require_POST
def eliminar_asignacion_ajax(request, pk):
    asignacion = get_object_or_404(AsignacionAcademica, pk=pk)
    asignacion.delete()
    return JsonResponse({'status': 'ok', 'message': 'Asignación eliminada correctamente.'})

# 1. Vista para mostrar la pantalla de Matrícula Masiva
def matricula_masiva(request):
    # Traemos todos los periodos activos para el select
    periodos = PeriodoLectivo.objects.filter(activo=True)
    # Usamos el primero para filtrar los estudiantes ya matriculados
    periodo_actual = periodos.first() 
    
    aulas = Aula.objects.all().order_by('nivel', 'grado', 'seccion')
    
    # Filtro de estudiantes no matriculados
    ids_ya_matriculados = Matricula.objects.filter(periodo=periodo_actual).values_list('estudiante_id', flat=True)
    estudiantes = Estudiante.objects.filter(estado='Activo').exclude(id__in=ids_ya_matriculados).order_by('apellidos')
    
    return render(request, 'academico/matricula_masiva.html', {
        'periodos': periodos,  # <--- Asegúrate de que se llame 'periodos'
        'aulas': aulas,
        'estudiantes': estudiantes
    })

# 2. Vista AJAX para procesar el guardado de muchos alumnos a la vez
@require_POST
def procesar_matricula_masiva(request):
    try:
        data = json.loads(request.body)
        periodo_id = data.get('periodo_id')
        aula_id = data.get('aula_id')
        estudiantes_ids = data.get('estudiantes_ids', [])

        if not periodo_id or not aula_id or not estudiantes_ids:
            return JsonResponse({'status': 'error', 'message': 'Faltan datos requeridos.'})

        periodo = get_object_or_404(PeriodoLectivo, pk=periodo_id)
        aula = get_object_or_404(Aula, pk=aula_id)

        # 💥 OPTIMIZACIÓN SENIOR (Bulk Operations)
        # 1. Buscamos a los que ya estaban matriculados y les cambiamos el aula rápido
        Matricula.objects.filter(
            periodo=periodo, 
            estudiante_id__in=estudiantes_ids
        ).update(aula=aula)
        
        # 2. Identificamos quiénes faltan por crear
        ya_matriculados = Matricula.objects.filter(
            periodo=periodo, estudiante_id__in=estudiantes_ids
        ).values_list('estudiante_id', flat=True)
        
        nuevos_ids = set(estudiantes_ids) - set(ya_matriculados)
        
        # 3. Creamos todas las matrículas nuevas de 1 solo golpe
        nuevas_matriculas = [
            Matricula(estudiante_id=eid, aula=aula, periodo=periodo) 
            for eid in nuevos_ids
        ]
        Matricula.objects.bulk_create(nuevas_matriculas)

        return JsonResponse({
            'status': 'ok', 
            'message': f'Se matricularon/actualizaron {len(estudiantes_ids)} estudiantes con éxito.'
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def lista_matriculas(request):
    # Traemos las matrículas del periodo activo para ver quién está en cada salón
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    matriculas = Matricula.objects.filter(periodo=periodo_actual).select_related('estudiante', 'aula')
    return render(request, 'academico/lista_matriculas.html', {'matriculas': matriculas})

@require_POST
def eliminar_matricula_ajax(request, pk):
    matricula = get_object_or_404(Matricula, pk=pk)
    nombre_alumno = f"{matricula.estudiante.nombres} {matricula.estudiante.apellidos}"
    matricula.delete()
    return JsonResponse({
        'status': 'ok', 
        'message': f'Se ha quitado a {nombre_alumno} de este salón correctamente.'
    })

def panel_asistente_imprenta(request):
    # Traemos todas las solicitudes, pero las PENDIENTES y EN PROCESO van primero
    solicitudes = SolicitudImpresion.objects.all().order_by('estado', '-fecha_subida')
    
    return render(request, 'asistente/panel_imprenta.html', {
        'solicitudes': solicitudes
    })

def actualizar_estado_impresion(request, solicitud_id):
    """Llamada AJAX para cambiar el estado sin recargar"""
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        solicitud = get_object_or_404(SolicitudImpresion, id=solicitud_id)
        solicitud.estado = nuevo_estado
        solicitud.save()
        
        # ==============================================================
        # 💥 MAGIA DOBLE: Notificación en el Sistema + WhatsApp
        # ==============================================================
        if nuevo_estado == 'LISTO':
            # Extraemos los datos básicos
            telefono_docente = solicitud.personal.telefono
            nombre = solicitud.personal.nombres.split()[0] # Primer nombre
            curso = solicitud.asignacion.curso.nombre
            tema = solicitud.get_tema_display()
            
            # 1. NOTIFICACIÓN IN-APP (Campanita del sistema)
            # Asumiendo que tu modelo 'Personal' tiene un campo de enlace al 'User' de Django
            if hasattr(solicitud.personal, 'usuario') and solicitud.personal.usuario:
                Notificacion.objects.create(
                    usuario=solicitud.personal.usuario,
                    titulo="¡Copias Listas! 🖨️",
                    mensaje=f"Sus materiales impresos para {curso} ({tema}) están listos en secretaría.",
                    tipo='EXITO',
                    # Opcional: El enlace a donde irá si le da clic en la campanita
                    enlace="/personal/centro_materiales/" 
                )
            
            # 2. NOTIFICACIÓN POR WHATSAPP (Usando nuestra Capa de Servicios)
            if telefono_docente:
                WhatsAppService.notificar_material_impreso(
                    nombre_docente=nombre,
                    telefono_docente=telefono_docente,
                    curso=curso,
                    tema=tema
                )
        # ==============================================================

        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

def obtener_archivos_solicitud(request, solicitud_id):
    """Devuelve los archivos de una solicitud en formato JSON para el Modal"""
    solicitud = get_object_or_404(SolicitudImpresion, id=solicitud_id)
    archivos_data = []
    
    for arc in solicitud.archivos.all():
        archivos_data.append({
            'nombre': arc.archivo.name.split('/')[-1], # Solo el nombre del archivo
            'tipo': arc.get_tipo_display(),
            'url': arc.archivo.url
        })
        
    return JsonResponse({'archivos': archivos_data})

@login_required
@never_cache
def mi_aula(request):
    personal_actual = obtener_personal_logueado(request)
    if not personal_actual:
        return render(request, 'errores/sin_perfil.html', {'mensaje': 'Sin perfil activo.'})
    
    aula_tutor = Aula.objects.filter(tutor=personal_actual).first() or Aula.objects.first()
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    cursos = Curso.objects.filter(
        asignaciones__aula=aula_tutor,
        asignaciones__periodo=periodo_actual,
        activo=True
    ).distinct()
    
    curso_id_param = request.GET.get('curso_id', '')
    curso_seleccionado = int(curso_id_param) if curso_id_param.isdigit() else None
    
    alumnos_procesados = []
    conteo_estados = {'danger': 0, 'warning': 0, 'success': 0}
    
    # 💥 OPTIMIZACIÓN SENIOR (N+1 RESUELTO): 
    # Traemos las matrículas, cruzamos con estudiante (select_related) 
    # y precargamos TODAS sus asistencias de un solo golpe (prefetch_related)
    matriculas = Matricula.objects.filter(
        aula=aula_tutor, 
        periodo=periodo_actual,
        estudiante__estado='Activo'
    ).select_related('estudiante').prefetch_related('estudiante__asistencias')
    
    for matricula in matriculas:
        estudiante = matricula.estudiante
        analisis_ia = {'promedio': 0, 'tendencia_numerica': 0, 'estado_ia': 'Sin Notas', 'color': 'secondary', 'cantidad_notas': 0}
        
        # 💥 OPTIMIZACIÓN: Como usamos prefetch_related, usar .all() lo saca de la memoria RAM, no de la BD
        asistencias = estudiante.asistencias.all()
        asistencias_totales = len(asistencias)
        
        if asistencias_totales > 0:
            # Filtramos en memoria (usando una lista de comprensión)
            asistencias_efectivas = len([a for a in asistencias if a.estado in ['P', 'T', 'J']])
            porcentaje_asistencia = int((asistencias_efectivas / asistencias_totales) * 100)
        else:
            porcentaje_asistencia = 100 

        # Tu función de IA
        analisis_ia = analizar_rendimiento_estudiante(matricula.id, curso_seleccionado)
        
        if analisis_ia['color'] in conteo_estados:
            conteo_estados[analisis_ia['color']] += 1
                
        alumnos_procesados.append({
            'objeto': estudiante,
            'ia': analisis_ia,
            'matricula_id': matricula.id,
            'asistencia': porcentaje_asistencia
        })
        
    return render(request, 'academico/mi_aula.html', {
        'aula': aula_tutor,
        'alumnos': alumnos_procesados,
        'totales': conteo_estados,
        'total_alumnos': matriculas.count(),
        'cursos': cursos,
        'curso_seleccionado': curso_seleccionado,
        'periodo': periodo_actual,
        'segment': 'mi_aula'
    })

@login_required
def generar_diagnostico_ajax(request):
    """
    Recibe la petición AJAX desde el JavaScript modular, extrae el alumno y el curso,
    y solicita a Gemini un análisis cualitativo adaptado.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            matricula_id = data.get('matricula_id')
            nombre_alumno = data.get('nombre_alumno')
            
            # 💥 NUEVO: Atrapamos el curso enviado por JavaScript
            curso_id_raw = data.get('curso_id', '')
            curso_id = int(curso_id_raw) if str(curso_id_raw).isdigit() else None

            # 1. Obtenemos la matemática exacta calculada (pasándole el curso_id)
            analisis_ia = analizar_rendimiento_estudiante(matricula_id, curso_id)

            # 2. Buscamos el nombre del curso si existe para darle contexto a la IA
            contexto_curso = "en general (todas las materias)"
            if curso_id:
                curso_obj = Curso.objects.filter(id=curso_id).first()
                if curso_obj:
                    contexto_curso = f"específicamente en el curso de {curso_obj.nombre}"

            # 3. Llamamos a tu modelo 'gemini-3.1-flash-lite' en servicios_ia.py
            diagnostico_texto = generar_diagnostico_cualitativo(
                nombre_alumno=nombre_alumno,
                promedio=analisis_ia['promedio'],
                tendencia=analisis_ia['tendencia_numerica'],
                estado_ia=analisis_ia['estado_ia'],
                contexto_curso=contexto_curso # Le pasamos qué curso es
            )

            return JsonResponse({'status': 'success', 'diagnostico': diagnostico_texto})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'mensaje': str(e)})
            
    return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido.'})

@require_POST
def generar_clustering_ia_api(request):
    try:
        data = json.loads(request.body)
        aula_id = data.get('aula_id')
        periodo_id = data.get('periodo_id') # Asumo que lo pasaremos desde el JS
        
        # Llamamos al algoritmo
        resultado = agrupar_estudiantes_kmeans(aula_id, periodo_id)
        
        if "error" in resultado:
            return JsonResponse({"status": "error", "mensaje": resultado["error"]})
            
        return JsonResponse(resultado)
        
    except Exception as e:
        print(f"Error de Machine Learning: {str(e)}")
        return JsonResponse({"status": "error", "mensaje": "Fallo al procesar el algoritmo de clustering."})
    
@login_required
def panel_supervision(request):
    personal_actual = obtener_personal_logueado(request)
    if not personal_actual:
        return render(request, 'errores/sin_perfil.html', {'mensaje': 'Sin perfil activo.'})

    # Asumimos que tienes un campo 'rol' o similar (ej: 'COORDINADOR', 'DOCENTE')
    # Corrección rápida en tu views.py dentro de panel_supervision:
    es_coordinador = personal_actual.cargo in ['COO', 'DIR'] or request.user.is_superuser

    materiales = MaterialInstitucional.objects.all()
    form_material = MaterialInstitucionalForm()
    form_evidencia = EvidenciaDocenteForm()

    if es_coordinador:
        # La coordinadora ve todas las evidencias de todos los docentes
        evidencias = EvidenciaDocente.objects.all().select_related('docente')
    else:
        # El docente solo ve sus propias entregas
        evidencias = EvidenciaDocente.objects.filter(docente=personal_actual)

    return render(request, 'academico/panel_supervision.html', {
        'es_coordinador': es_coordinador,
        'materiales': materiales,
        'evidencias': evidencias,
        'form_material': form_material,
        'form_evidencia': form_evidencia,
        'personal': personal_actual,
        'segment': 'supervision'
    })

@login_required
@require_POST
def guardar_material_ajax(request):
    """ API para que la Coordinadora suba un nuevo formato/guía """
    personal_actual = obtener_personal_logueado(request)
    form = MaterialInstitucionalForm(request.POST, request.FILES)
    if form.is_valid():
        material = form.save(commit=False)
        material.autor = personal_actual
        material.save()
        return JsonResponse({'success': True, 'mensaje': 'Material institucional publicado.'})
    return JsonResponse({'success': False, 'mensaje': 'Datos inválidos.', 'errors': form.errors})

@login_required
@require_POST
def eliminar_material_ajax(request, pk):
    """ API para eliminar un formato del repositorio """
    material = get_object_or_404(MaterialInstitucional, pk=pk)
    material.archivo.delete() # Borra el archivo físico en la carpeta media
    material.delete()
    return JsonResponse({'success': True, 'mensaje': 'Material eliminado correctamente.'})

@login_required
@require_POST
def guardar_evidencia_ajax(request):
    """ API para que los Docentes suban sus evidencias (fichas, paneles, etc.) """
    personal_actual = obtener_personal_logueado(request)
    form = EvidenciaDocenteForm(request.POST, request.FILES)
    
    if form.is_valid():
        evidencia = form.save(commit=False)
        evidencia.docente = personal_actual
        evidencia.estado = 'PENDIENTE'
        evidencia.save()
        
        # ==============================================================
        # 💥 MAGIA: Notificamos a la(s) Coordinadora(s) (Solo Campanita)
        # ==============================================================
        # Buscamos en tu modelo Personal a todas las que sean 'COO' y estén activas
        coordinadoras = Personal.objects.filter(cargo='COO', estado='Activo').exclude(user__isnull=True)
        
        for coord in coordinadoras:
            Notificacion.objects.create(
                usuario=coord.user,  # Usamos el campo 'user' de tu modelo Personal
                titulo="Nuevos Materiales Subidos 📁",
                mensaje=f"El Prof. {personal_actual.nombres.split()[0]} ha subido una nueva evidencia para revisión.",
                tipo='INFO',
                enlace="/academico/gestion-evidencias/" # Ajusta a la URL real
            )
        # ==============================================================

        return JsonResponse({'success': True, 'mensaje': 'Evidencia enviada a coordinación con éxito.'})
    
    return JsonResponse({'success': False, 'mensaje': 'Asegúrate de adjuntar un archivo válido.'})


@login_required
@require_POST
def revisar_evidencia_ajax(request):
    """ API para que la Coordinadora Apruebe u Observe una evidencia """
    personal_actual = obtener_personal_logueado(request)
    
    evidencia_id = request.POST.get('id')
    nuevo_estado = request.POST.get('estado')
    observaciones = request.POST.get('observaciones_coordinacion', '')

    evidencia = get_object_or_404(EvidenciaDocente, id=evidencia_id)
    evidencia.estado = nuevo_estado
    evidencia.observaciones_coordinacion = observaciones
    evidencia.fecha_revision = localtime(now()).date()
    evidencia.revisado_por = personal_actual
    evidencia.save()

    # ==============================================================
    # 💥 MAGIA DOBLE: Notificamos al Docente (Campanita + WhatsApp)
    # ==============================================================
    docente = evidencia.docente # Esto es una instancia de tu modelo Personal
    
    # 1. Campanita roja (Revisamos que tenga una cuenta de Django enlazada)
    if docente.user:
        tipo_noti = 'EXITO' if nuevo_estado == 'APROBADO' else 'ALERTA'
        Notificacion.objects.create(
            usuario=docente.user, # Usamos 'user' en lugar de 'usuario'
            titulo=f"Evidencia {evidencia.get_estado_display()}",
            mensaje=f"Su evidencia ha sido revisada por coordinación.",
            tipo=tipo_noti,
            enlace="javascript:;" # Ajusta a la URL donde el profe ve sus evidencias
        )

    # 2. WhatsApp Inmediato (Revisamos que tenga teléfono registrado)
    if docente.telefono:
        nombre = docente.nombres.split()[0]
        WhatsAppService.notificar_revision_evidencia(
            nombre_docente=nombre,
            telefono_docente=docente.telefono,
            estado=evidencia.get_estado_display(),
            observaciones=observaciones
        )
    # ==============================================================

    return JsonResponse({'success': True, 'mensaje': f'Evidencia marcada como {evidencia.get_estado_display()}.'})

# ==========================================
# 1. MÓDULO DE AGENDA Y CRONOGRAMA DINÁMICO
# ==========================================

@login_required
def panel_cronograma(request):
    personal_actual = obtener_personal_logueado(request)
    if not personal_actual:
        return render(request, 'errores/sin_perfil.html', {'mensaje': 'Sin perfil activo.'})

    es_coordinador = personal_actual.cargo in ['COO', 'DIR'] or request.user.is_superuser
    
    form_horario = HorarioClaseForm()
    form_evento = EventoCronogramaForm()

    hoy = localtime(now()).date()
    dias_codigo = {0: 'LU', 1: 'MA', 2: 'MI', 3: 'JU', 4: 'VI', 5: 'SA', 6: 'DO'}
    codigo_hoy = dias_codigo.get(hoy.weekday(), 'LU')

    if es_coordinador:
        horarios = HorarioClase.objects.all().select_related('personal', 'aula', 'curso')
        eventos = EventoCronograma.objects.all()
    else:
        horarios = HorarioClase.objects.filter(
            personal=personal_actual, 
            dia_semana=codigo_hoy
        ).select_related('aula', 'curso').order_by('hora_inicio')
        
        eventos = EventoCronograma.objects.filter(fecha_inicio__gte=hoy)

    return render(request, 'academico/panel_cronograma.html', {
        'es_coordinador': es_coordinador,
        'horarios': horarios,
        'eventos': eventos,
        'form_horario': form_horario,
        'form_evento': form_evento,
        'personal': personal_actual,
        'segment': 'cronograma'
    })

@login_required
def api_calendario_eventos(request):
    """ Endpoint JSON para la Agenda Dinámica """
    personal_actual = obtener_personal_logueado(request)
    es_coordinador = personal_actual.cargo in ['COO', 'DIR'] or request.user.is_superuser
    eventos_fc = []

    agenda = EventoCronograma.objects.all() if es_coordinador else EventoCronograma.objects.filter(fecha_fin__gte=localtime(now()).date())
    
    for e in agenda:
        start_iso = f"{e.fecha_inicio.strftime('%Y-%m-%d')}T{e.hora_inicio.strftime('%H:%M:%S')}" if e.hora_inicio else e.fecha_inicio.strftime('%Y-%m-%d')
        
        if e.hora_fin:
            end_iso = f"{e.fecha_fin.strftime('%Y-%m-%d')}T{e.hora_fin.strftime('%H:%M:%S')}"
        else:
            ff_ajustada = e.fecha_fin + datetime.timedelta(days=1)
            end_iso = ff_ajustada.strftime('%Y-%m-%d')
        
        eventos_fc.append({
            'id': f'evt_{e.id}',
            'title': e.titulo,
            'start': start_iso,
            'end': end_iso,
            'backgroundColor': e.color,
            'borderColor': e.color,
            'allDay': False if e.hora_inicio else True,
            'extendedProps': {
                'tipo': 'evento',
                'db_id': e.id,
                'descripcion': e.descripcion or '',
                'fecha_inicio_raw': e.fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_fin_raw': e.fecha_fin.strftime('%Y-%m-%d'),
                'hora_inicio_raw': e.hora_inicio.strftime('%H:%M') if e.hora_inicio else '',
                'hora_fin_raw': e.hora_fin.strftime('%H:%M') if e.hora_fin else '',
                'color_raw': e.color,
                'aula_id': e.aula_afectada.id if e.aula_afectada else ''
            }
        })

    if not es_coordinador:
        clases = HorarioClase.objects.filter(personal=personal_actual).select_related('aula', 'curso')
        mapa_dias = {'DO': 0, 'LU': 1, 'MA': 2, 'MI': 3, 'JU': 4, 'VI': 5, 'SA': 6}
        for c in clases:
            eventos_fc.append({
                'id': f'cls_{c.id}',
                'title': f"{c.curso.nombre} ({c.aula.grado} '{c.aula.seccion}')",
                'startTime': c.hora_inicio.strftime('%H:%M:%S'),
                'endTime': c.hora_fin.strftime('%H:%M:%S'),
                'daysOfWeek': [mapa_dias[c.dia_semana]],
                'backgroundColor': '#17c1e8',
                'borderColor': '#17c1e8',
                'extendedProps': {
                    'tipo': 'clase',
                    'descripcion': f'Aula: {c.aula.grado} "{c.aula.seccion}"'
                }
            })

    return JsonResponse(eventos_fc, safe=False)

@login_required
@require_POST
def guardar_evento_ajax(request):
    evento_id = request.POST.get('evento_id')
    
    if evento_id:
        evento = get_object_or_404(EventoCronograma, id=evento_id)
        form = EventoCronogramaForm(request.POST, instance=evento)
        mensaje = "Evento actualizado correctamente."
    else:
        form = EventoCronogramaForm(request.POST)
        mensaje = "Evento publicado en la agenda."

    if form.is_valid():
        try:
            evento = form.save(commit=False)
            if not evento_id:
                evento.creado_por = obtener_personal_logueado(request)
            
            hora_ini = request.POST.get('hora_inicio')
            hora_f = request.POST.get('hora_fin')
            
            evento.hora_inicio = hora_ini if hora_ini else None
            evento.hora_fin = hora_f if hora_f else None
            
            evento.save()
            return JsonResponse({'success': True, 'mensaje': mensaje})
        except Exception as e:
            return JsonResponse({'success': False, 'mensaje': str(e)})
            
    return JsonResponse({'success': False, 'mensaje': 'Revisa los campos cargados.', 'errors': form.errors.as_json()})

@login_required
@require_POST
def eliminar_evento_ajax(request, pk):
    evento = get_object_or_404(EventoCronograma, pk=pk)
    evento.delete()
    return JsonResponse({'success': True, 'mensaje': 'Evento eliminado del calendario.'})

@login_required
@require_POST
def actualizar_evento_drag_ajax(request):
    evento_id = request.POST.get('evento_id')
    fecha_inicio = request.POST.get('fecha_inicio')
    fecha_fin = request.POST.get('fecha_fin')
    hora_inicio = request.POST.get('hora_inicio')
    hora_fin = request.POST.get('hora_fin')

    try:
        evento = EventoCronograma.objects.get(id=evento_id)
        evento.fecha_inicio = fecha_inicio
        evento.fecha_fin = fecha_fin
        
        evento.hora_inicio = hora_inicio if hora_inicio else None
        evento.hora_fin = hora_fin if hora_fin else None
        
        evento.save()
        return JsonResponse({'success': True, 'mensaje': 'Agenda reprogramada automáticamente.'})
    except Exception as e:
        return JsonResponse({'success': False, 'mensaje': 'Error al actualizar el evento.'})

# ==========================================
# 2. MÓDULO DE HORARIO ESCOLAR (GENERAL Y AULA)
# ==========================================

@login_required
def horario_maestro(request):
    """ Panel principal del Horario Fijo (Vista de Lectura y Filtros) """
    personal_actual = obtener_personal_logueado(request)
    cargo = personal_actual.cargo
    
    # 1. Definición estricta de Roles
    es_coordinador = cargo in ['COO', 'DIR'] or request.user.is_superuser
    puede_ver_todo = cargo in ['COO', 'DIR', 'SEC', 'ASI'] or request.user.is_superuser
    
    # 2. Filtrado de Datos según el Rol
    if puede_ver_todo:
        docentes = Personal.objects.filter(cargo='DOC', estado='Activo')
        aulas = Aula.objects.all().order_by('nivel', 'grado', 'seccion')
    else:
        # 💥 Perfil Docente: Solo se ve a sí mismo
        docentes = Personal.objects.filter(id=personal_actual.id)
        # 💥 Solo ve las aulas donde tiene Asignaciones Académicas (evita duplicados con distinct)
        aulas = Aula.objects.filter(
            asignaciones__personal=personal_actual
        ).distinct().order_by('nivel', 'grado', 'seccion')

    periodos = PeriodoLectivo.objects.all()
    form_horario = HorarioClaseForm()
    
    return render(request, 'academico/horario_maestro.html', {
        'es_coordinador': es_coordinador,
        'puede_ver_todo': puede_ver_todo, # Enviamos esta variable clave al HTML
        'docentes': docentes,
        'aulas': aulas,
        'periodos': periodos,
        'form_horario': form_horario,
        'segment': 'horario_maestro'
    })

@login_required
def gestionar_horario_aula(request, aula_id):
    """ FUSIÓN CORREGIDA: Pantalla independiente con Drag & Drop exclusivo para un Aula """
    personal_actual = obtener_personal_logueado(request)
    es_coordinador = personal_actual.cargo in ['COO', 'DIR'] or request.user.is_superuser
    
    if not es_coordinador:
        return redirect('horario_maestro') # Seguridad
        
    aula = get_object_or_404(Aula, id=aula_id)
    periodos = PeriodoLectivo.objects.all()
    docentes = Personal.objects.filter(cargo='DOC', estado='Activo')
    
    form_horario = HorarioClaseForm(initial={'aula': aula})
    
    return render(request, 'academico/gestionar_horario_aula.html', {
        'aula': aula,
        'periodos': periodos,
        'docentes': docentes,
        'form_horario': form_horario,
        'es_coordinador': es_coordinador,
        'segment': 'horario_maestro'
    })

@login_required
def api_horario_fijo(request):
    """ API JSON optimizada: Carga clases regulares y bloqueos de Recreo Globales """
    periodo_id = request.GET.get('periodo_id')
    nivel = request.GET.get('nivel')
    aula_id = request.GET.get('aula_id')
    docente_id = request.GET.get('docente_id')
    
    clases = HorarioClase.objects.all().select_related('aula', 'curso', 'personal')
    
    if periodo_id: clases = clases.filter(periodo_id=periodo_id)
    if nivel: clases = clases.filter(aula__nivel=nivel)
    if aula_id: clases = clases.filter(aula_id=aula_id)
    if docente_id: clases = clases.filter(personal_id=docente_id)
        
    # Semana ficticia ancla de lunes a viernes
    anclas = {'LU': '2024-01-01', 'MA': '2024-01-02', 'MI': '2024-01-03', 'JU': '2024-01-04', 'VI': '2024-01-05'}
    
    eventos_fc = []
    
    # =========================================================
    # 💥 INYECCIÓN DE RECREOS COMO EVENTOS DE FONDO (BLOQUEANTES)
    # =========================================================
    niveles_a_cargar = []
    if nivel:
        niveles_a_cargar.append(nivel)
    elif aula_id:
        aula_obj = Aula.objects.filter(id=aula_id).first()
        if aula_obj:
            niveles_a_cargar.append(aula_obj.nivel)
    else:
        # Si no hay filtros cruzados, cargamos todos para servir de guía visual global
        niveles_a_cargar = ['INICIAL', 'PRIMARIA', 'SECUNDARIA']
        
    recreos = HorarioRecreo.objects.filter(nivel__in=niveles_a_cargar)
    for r in recreos:
        for dia_cod, fecha_str in anclas.items():
            eventos_fc.append({
                'id': f"rec_{r.id}_{dia_cod}",
                'title': r.nombre,
                'start': f"{fecha_str}T{r.hora_inicio.strftime('%H:%M:%S')}",
                'end': f"{fecha_str}T{r.hora_fin.strftime('%H:%M:%S')}",
                'display': 'background',          # 💥 Pinta el bloque horizontal de fondo
                'backgroundColor': "#BEBEBE",     # Color naranja pastel muy sutil estilo Material
                'overlap': False,                 # 💥 IMPEDIR: Bloquea arrastres y clics encima de esta celda
                'extendedProps': { 'tipo': 'recreo' }
            })
    # =========================================================

    # Carga regular de las clases académicas
    for c in clases:
        fecha_base = anclas.get(c.dia_semana, '2024-01-01')
        color = c.color if hasattr(c, 'color') and c.color else '#1a73e8'
        
        eventos_fc.append({
            'id': str(c.id),
            'title': f"{c.curso.nombre}",
            'start': f"{fecha_base}T{c.hora_inicio.strftime('%H:%M:%S')}",
            'end': f"{fecha_base}T{c.hora_fin.strftime('%H:%M:%S')}",
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'db_id': c.id,
                'curso': c.curso.nombre,
                'curso_id': c.curso.id,
                'personal_id': c.personal.id,
                'aula': f"{c.aula.grado} '{c.aula.seccion}'",
                'nivel': c.aula.get_nivel_display(),
                'docente': f"{c.personal.apellidos}, {c.personal.nombres}",
                'dia_semana': c.dia_semana,
                'color_raw': color,
                'hora_inicio_txt': c.hora_inicio.strftime('%I:%M %p'),
                'hora_fin_txt': c.hora_fin.strftime('%I:%M %p'),
            }
        })
        
    return JsonResponse(eventos_fc, safe=False)

@login_required
@require_POST
def guardar_horario_ajax(request):
    """ API unificada para guardar (CREAR) y actualizar (EDITAR) un bloque de clase """
    horario_id = request.POST.get('horario_id') # <--- Capturamos el ID del modal de JS
    
    if horario_id:
        # Si el ID existe, estamos EDITANDO: recuperamos el registro y se lo pasamos al Form
        horario = get_object_or_404(HorarioClase, id=horario_id)
        form = HorarioClaseForm(request.POST, instance=horario)
        mensaje = 'Bloque de horario actualizado con éxito.'
    else:
        # Si no hay ID, es un casillero vacío: se crea un registro NUEVO
        form = HorarioClaseForm(request.POST)
        mensaje = 'Bloque de horario asignado con éxito.'
        
    if form.is_valid():
        try:
            form.save()
            return JsonResponse({'success': True, 'mensaje': mensaje})
        except Exception as e:
            return JsonResponse({'success': False, 'mensaje': str(e)})
    
    # Si falla la validación (cruces en el clean del modelo, etc.)
    errores = form.errors.as_json()
    return JsonResponse({'success': False, 'mensaje': 'Cruce de horarios detectado o datos inválidos.', 'errors': errores})

@login_required
@require_POST
def actualizar_horario_drag_ajax(request):
    """ API que procesa el arrastre y soltado de clases en vivo para el Horario Fijo """
    horario_id = request.POST.get('horario_id')
    nuevo_dia = request.POST.get('dia_semana')
    hora_inicio = request.POST.get('hora_inicio')
    hora_fin = request.POST.get('hora_fin')

    try:
        horario = HorarioClase.objects.get(id=horario_id)
        horario.dia_semana = nuevo_dia
        horario.hora_inicio = hora_inicio
        horario.hora_fin = hora_fin
        
        horario.clean() 
        horario.save()
        return JsonResponse({'success': True, 'mensaje': 'Horario actualizado.'})
    except ValidationError as e:
        mensaje = e.message if hasattr(e, 'message') else str(e.messages[0] if hasattr(e, 'messages') else e)
        return JsonResponse({'success': False, 'mensaje': mensaje})
    except Exception as e:
        return JsonResponse({'success': False, 'mensaje': 'Error de servidor.'})

@login_required
@require_POST
def eliminar_horario_ajax(request, pk):
    """ API para borrar un bloque de horario """
    horario = get_object_or_404(HorarioClase, pk=pk)
    horario.delete()
    return JsonResponse({'success': True, 'mensaje': 'Bloque de clase eliminado.'})

@login_required
def api_cursos_docente(request, docente_id):
    """ API que devuelve solo los cursos asignados a un docente (filtrado por aula y periodo si se envían) """
    
    # Capturamos los filtros extra que enviaremos desde el JavaScript
    aula_id = request.GET.get('aula_id')
    periodo_id = request.GET.get('periodo_id')
    
    # 💥 MAGIA DEL ORM: Buscamos cursos que tengan una asignación con este docente
    # Usamos distinct() por si dicta el mismo curso en varias aulas y no enviaron filtro
    cursos_query = Curso.objects.filter(asignaciones__personal_id=docente_id, activo=True)
    
    # Si estamos en el editor de un aula específica, somos más estrictos con el filtro:
    if aula_id:
        cursos_query = cursos_query.filter(asignaciones__aula_id=aula_id)
    if periodo_id:
        cursos_query = cursos_query.filter(asignaciones__periodo_id=periodo_id)
        
    cursos = cursos_query.distinct().values('id', 'nombre')
    
    return JsonResponse(list(cursos), safe=False)

@login_required
def exportar_horario_excel(request):
    """ Genera un archivo Excel estructurado como grilla escolar según los filtros activos """
    periodo_id = request.GET.get('periodo_id')
    aula_id = request.GET.get('aula_id')
    docente_id = request.GET.get('docente_id')

    clases = HorarioClase.objects.all().select_related('aula', 'curso', 'personal')
    titulo_reporte = "Matriz General de Horarios Escolares"
    
    # 1. Filtros y Títulos
    aula = None
    if periodo_id:
        clases = clases.filter(periodo_id=periodo_id)
    if aula_id:
        clases = clases.filter(aula_id=aula_id)
        aula = Aula.objects.filter(id=aula_id).first()
        if aula:
            titulo_reporte = f"Horario Escolar: {aula.grado} '{aula.seccion}' — {aula.get_nivel_display()}"
    if docente_id:
        clases = clases.filter(personal_id=docente_id)
        docente = Personal.objects.filter(id=docente_id).first()
        if docente:
            titulo_reporte = f"Horario Semanal: {docente.apellidos}, {docente.nombres}"

    # 2. Capturar los Recreos según el contexto
    niveles_a_cargar = []
    if aula:
        niveles_a_cargar.append(aula.nivel)
    elif docente_id:
        # Si filtramos por docente, traemos los recreos de los niveles en los que dicta clase
        niveles_a_cargar = list(clases.values_list('aula__nivel', flat=True).distinct())
    else:
        niveles_a_cargar = ['INICIAL', 'PRIMARIA', 'SECUNDARIA']
        
    recreos = HorarioRecreo.objects.filter(nivel__in=niveles_a_cargar)

    # 3. UNIR HORAS DE CLASE Y RECREO (Evita que la fila de recreo desaparezca)
    intervalos_set = set()
    for c in clases.values('hora_inicio', 'hora_fin').distinct():
        intervalos_set.add((c['hora_inicio'], c['hora_fin']))
    for r in recreos.values('hora_inicio', 'hora_fin').distinct():
        intervalos_set.add((r['hora_inicio'], r['hora_fin']))
        
    # Ordenamos cronológicamente las horas de menor a mayor
    intervalos_ordenados = sorted(list(intervalos_set), key=lambda x: x[0])
    
    # 4. Inicializar el Libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horario Oficial"
    ws.views.sheetView[0].showGridLines = True 

    # 5. Definición de Estilos Premium 
    fill_titulo = PatternFill(start_color="f57c00", end_color="f57c00", fill_type="solid") 
    fill_cabecera = PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid") 
    fill_celda_hora = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") 
    
    # 💥 Estilos exclusivos para el Recreo
    fill_recreo = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid") 
    font_recreo = Font(name="Arial", size=12, bold=True, color="E65100") 
    
    font_titulo = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    font_cabecera = Font(name="Arial", size=11, bold=True, color="000000")
    font_cuerpo = Font(name="Arial", size=10, bold=False, color="000000")
    font_hora = Font(name="Arial", size=10, bold=True, color="000000")
    
    align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_fino = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )

    # 6. Inyectar Banner de Título Superior
    ws.merge_cells("A1:F1")
    ws["A1"] = titulo_reporte
    ws["A1"].font = font_titulo
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = align_centro
    ws.row_dimensions[1].height = 40

    # 7. Configurar Cabeceras de la Grilla (Columnas)
    columnas_dias = [("A", "HORAS"), ("B", "LUNES"), ("C", "MARTES"), ("D", "MIÉRCOLES"), ("E", "JUEVES"), ("F", "VIERNES")]
    ws.row_dimensions[3].height = 25
    for col, nombre in columnas_dias:
        celda = f"{col}3"
        ws[celda] = nombre
        ws[celda].font = font_cabecera
        ws[celda].fill = fill_cabecera
        ws[celda].alignment = align_centro
        ws[celda].border = border_fino

    mapeo_columnas = {'LU': 'B', 'MA': 'C', 'MI': 'D', 'JU': 'E', 'VI': 'F'}

    # 8. Construcción Dinámica de la Matriz (Filas)
    fila_actual = 4
    for h_ini, h_fin in intervalos_ordenados:
        
        # Escribir el bloque de hora en la columna A
        rango_texto = f"{h_ini.strftime('%I:%M %p')} - {h_fin.strftime('%I:%M %p')}"
        ws[f"A{fila_actual}"] = rango_texto
        ws[f"A{fila_actual}"].font = font_hora
        ws[f"A{fila_actual}"].fill = fill_celda_hora
        ws[f"A{fila_actual}"].alignment = align_centro
        ws[f"A{fila_actual}"].border = border_fino
        
        # 💥 VERIFICAMOS SI ESTA FILA ES UN RECREO
        recreo_en_rango = recreos.filter(hora_inicio=h_ini, hora_fin=h_fin).first()

        if recreo_en_rango:
            # COMBINAR CELDAS DE LUNES A VIERNES (Columnas 2 a 6)
            ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=6)
            
            celda_recreo = f"B{fila_actual}"
            texto_recreo = f"R E C R E O   {recreo_en_rango.get_nivel_display().upper() if not aula_id else ''}".strip()
            
            ws[celda_recreo] = texto_recreo
            ws[celda_recreo].font = font_recreo
            ws[celda_recreo].fill = fill_recreo
            ws[celda_recreo].alignment = align_centro
            
            # Pintar bordes en todas las celdas combinadas para que no se vea roto
            for col_letra in ['B', 'C', 'D', 'E', 'F']:
                ws[f"{col_letra}{fila_actual}"].border = border_fino
                
        else:
            # SI NO ES RECREO, ES UNA FILA DE CLASES NORMAL
            for col_letra in ['B', 'C', 'D', 'E', 'F']:
                ws[f"{col_letra}{fila_actual}"].border = border_fino
                ws[f"{col_letra}{fila_actual}"].alignment = align_centro
                ws[f"{col_letra}{fila_actual}"].font = font_cuerpo

            clases_en_rango = clases.filter(hora_inicio=h_ini, hora_fin=h_fin)
            for clase in clases_en_rango:
                col_letra = mapeo_columnas.get(clase.dia_semana)
                if col_letra:
                    if docente_id:
                        texto_celda = f"{clase.curso.nombre}\nAula: {clase.aula.grado} '{clase.aula.seccion}'"
                    else:
                        texto_celda = f"{clase.curso.nombre}\nProf: {clase.personal.apellidos}"
                    ws[f"{col_letra}{fila_actual}"] = texto_celda

        ws.row_dimensions[fila_actual].height = 45
        fila_actual += 1

    # 9. Auto-ajustar el ancho de las columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue 
            if cell.value:
                lineas = str(cell.value).split('\n')
                for l in lineas:
                    if len(l) > max_len: max_len = len(l)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

    # 10. Descarga del Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Horario_Escolar_Oficial.xlsx'
    wb.save(response)
    
    return response

@login_required
@require_POST
def guardar_recreo_ajax(request):
    """ API para configurar el horario de recreo por nivel """
    nivel = request.POST.get('nivel')
    hora_inicio = request.POST.get('hora_inicio')
    hora_fin = request.POST.get('hora_fin')

    if not all([nivel, hora_inicio, hora_fin]):
        return JsonResponse({'success': False, 'mensaje': 'Todos los campos son obligatorios.'})

    try:
        # Actualiza el recreo si ya existe para ese nivel, o lo crea si no existe
        recreo, created = HorarioRecreo.objects.update_or_create(
            nivel=nivel,
            defaults={
                'hora_inicio': hora_inicio,
                'hora_fin': hora_fin,
                'nombre': 'RECREO'
            }
        )
        accion = "creado" if created else "actualizado"
        return JsonResponse({'success': True, 'mensaje': f'Recreo de {nivel} {accion} con éxito.'})
    except Exception as e:
        return JsonResponse({'success': False, 'mensaje': 'Error en el servidor: ' + str(e)})

@login_required
def obtener_recreo_nivel_ajax(request, nivel):
    """ Devuelve las horas del recreo configurado para un nivel específico """
    recreo = HorarioRecreo.objects.filter(nivel=nivel).first()
    if recreo:
        return JsonResponse({
            'success': True,
            'hora_inicio': recreo.hora_inicio.strftime('%H:%M'),
            'hora_fin': recreo.hora_fin.strftime('%H:%M')
        })
    return JsonResponse({'success': False, 'mensaje': 'Sin configuración previa.'})

def gestionar_inventario_aula(request, aula_id):
    """ Vista interactiva para que el docente llene el inventario de su aula """
    aula = get_object_or_404(Aula, id=aula_id)
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    materiales_catalogo = CatalogoMaterial.objects.filter(activo=True)

    if not periodo_actual:
        messages.error(request, "No hay un periodo lectivo activo configurado.")
        return redirect('alguna_ruta_segura') # Cambia esto por tu ruta de inicio

    if request.method == 'POST':
        for material in materiales_catalogo:
            buen_estado = request.POST.get(f'bueno_{material.id}', 0)
            regular = request.POST.get(f'regular_{material.id}', 0)
            mal_estado = request.POST.get(f'malo_{material.id}', 0)
            se_requiere = request.POST.get(f'requiere_{material.id}', 0)

            InventarioAula.objects.update_or_create(
                periodo=periodo_actual,
                aula=aula,
                material=material,
                defaults={
                    'buen_estado': int(buen_estado) if buen_estado else 0,
                    'regular': int(regular) if regular else 0,
                    'mal_estado': int(mal_estado) if mal_estado else 0,
                    'se_requiere': int(se_requiere) if se_requiere else 0,
                }
            )
        
        # 💥 Si es una petición AJAX, respondemos con JSON en lugar de recargar la página
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'mensaje': 'Inventario guardado correctamente.'})
            
        # (Esto se queda por si acaso falla el JS, actúa como plan de respaldo)
        messages.success(request, "¡Inventario guardado exitosamente!")
        return redirect('academico:gestionar_inventario_aula', aula_id=aula.id)

    # ================= PARA LA VISUALIZACIÓN (MÉTODO GET) =================
    # Recuperamos lo que ya llenó antes (si existe) para mostrárselo
    inventario_previo = InventarioAula.objects.filter(periodo=periodo_actual, aula=aula)
    dict_inventario = {inv.material_id: inv for inv in inventario_previo}

    # Armamos una lista combinada lista para ser renderizada en el HTML
    datos_grilla = []
    for m in materiales_catalogo:
        inv = dict_inventario.get(m.id)
        datos_grilla.append({
            'id_material': m.id,
            'nombre': m.nombre,
            'bueno': inv.buen_estado if inv else 0,
            'regular': inv.regular if inv else 0,
            'malo': inv.mal_estado if inv else 0,
            'requiere': inv.se_requiere if inv else 0,
        })

    return render(request, 'academico/inventario_docente.html', {
        'aula': aula,
        'periodo': periodo_actual,
        'datos_grilla': datos_grilla
    })

def inventario_general(request):
    """ Vista para la Coordinadora: Muestra la suma total de todo el colegio """
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    
    if not periodo_actual:
        messages.error(request, "No hay un periodo lectivo activo configurado.")
        # Corrección: El redirect usa el 'name' de la URL, no la ruta del template
        return redirect('academico:lista_periodos') 

    # 💥 LA SOLUCIÓN: Consultamos todas las aulas para el select
    todas_las_aulas = Aula.objects.all().order_by('nivel', 'grado', 'seccion')

    # MAGIA DE BASE DE DATOS: Sumamos todos los salones en una sola consulta
    materiales_consolidados = CatalogoMaterial.objects.filter(activo=True).annotate(
        total_bueno=Coalesce(Sum('inventarios_aula__buen_estado', filter=Q(inventarios_aula__periodo=periodo_actual)), 0),
        total_regular=Coalesce(Sum('inventarios_aula__regular', filter=Q(inventarios_aula__periodo=periodo_actual)), 0),
        total_malo=Coalesce(Sum('inventarios_aula__mal_estado', filter=Q(inventarios_aula__periodo=periodo_actual)), 0),
        total_requiere=Coalesce(Sum('inventarios_aula__se_requiere', filter=Q(inventarios_aula__periodo=periodo_actual)), 0)
    )

    return render(request, 'academico/inventario_general.html', {
        'materiales': materiales_consolidados,
        'periodo': periodo_actual,
        'todas_las_aulas': todas_las_aulas  # 💥 AHORA SÍ PASAMOS LA VARIABLE AL HTML
    })

@login_required
def api_inventario_general(request):
    """ API JSON que alimenta el DataTables de la Coordinadora con soporte de filtros por aula """
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    
    # 1. Atrapamos el parámetro que nos envía el JavaScript de DataTables
    aula_id = request.GET.get('aula_id')
    
    # 2. Construimos la "Base" del filtro (Siempre debe ser del periodo actual)
    filtro_inventario = Q(inventarios_aula__periodo=periodo_actual)
    
    # 3. Si se seleccionó un aula en el HTML, le agregamos esa restricción al filtro
    if aula_id:
        filtro_inventario &= Q(inventarios_aula__aula_id=aula_id)
        
    # 4. Hacemos las sumas usando nuestro filtro dinámico
    materiales = CatalogoMaterial.objects.filter(activo=True).annotate(
        t_bueno=Coalesce(Sum('inventarios_aula__buen_estado', filter=filtro_inventario), 0),
        t_regular=Coalesce(Sum('inventarios_aula__regular', filter=filtro_inventario), 0),
        t_malo=Coalesce(Sum('inventarios_aula__mal_estado', filter=filtro_inventario), 0),
        t_req=Coalesce(Sum('inventarios_aula__se_requiere', filter=filtro_inventario), 0)
    )
    
    # 5. Formateamos la respuesta para DataTables
    data = []
    for m in materiales:
        data.append({
            'id': m.id,
            'nombre': m.nombre,
            'bueno': m.t_bueno,
            'regular': m.t_regular,
            'malo': m.t_malo,
            'requerido': m.t_req
        })
        
    return JsonResponse({'data': data})

@login_required
def exportar_inventario_excel(request):
    """ Exporta el inventario a Excel (General o por Aula específica) """
    aula_id = request.GET.get('aula_id')
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    # Estilos Premium
    fill_cabecera = PatternFill(start_color="344767", end_color="344767", fill_type="solid")
    font_cabecera = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_cuerpo = Font(name="Arial", size=10)
    align_centro = Alignment(horizontal="center", vertical="center")
    align_izq = Alignment(horizontal="left", vertical="center")
    borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    if aula_id:
        # EXPORTACIÓN ESPECÍFICA DE UN AULA
        aula = get_object_or_404(Aula, id=aula_id)
        ws.title = f"Inv. {aula.grado} {aula.seccion}"
        titulo = f"INVENTARIO FÍSICO DE AULA: {aula.grado} '{aula.seccion}' - {aula.get_nivel_display()}"
        
        datos = InventarioAula.objects.filter(periodo=periodo_actual, aula=aula).select_related('material')
        filas_datos = [(d.material.nombre, d.buen_estado, d.regular, d.mal_estado, d.se_requiere) for d in datos]
    else:
        # EXPORTACIÓN DEL CONSOLIDADO GLOBAL
        ws.title = "Consolidado General"
        titulo = f"CONSOLIDADO GLOBAL DE INVENTARIO - PERIODO {periodo_actual}"
        
        materiales = CatalogoMaterial.objects.filter(activo=True).annotate(
            t_bueno=Coalesce(Sum('inventarios_aula__buen_estado', filter=Q(inventarios_aula__periodo=periodo_actual)), 0),
            t_regular=Coalesce(Sum('inventarios_aula__regular', filter=Q(inventarios_aula__periodo=periodo_actual)), 0),
            t_malo=Coalesce(Sum('inventarios_aula__mal_estado', filter=Q(inventarios_aula__periodo=periodo_actual)), 0),
            t_req=Coalesce(Sum('inventarios_aula__se_requiere', filter=Q(inventarios_aula__periodo=periodo_actual)), 0)
        )
        filas_datos = [(m.nombre, m.t_bueno, m.t_regular, m.t_malo, m.t_req) for m in materiales]

    # Imprimir Título
    ws.merge_cells("A1:F1")
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = align_centro
    
    # ==========================================
    # IMPRIMIR CABECERAS DOBLES (TWO-TIER HEADERS)
    # ==========================================
    
    # 1. Combinar celdas verticales (Filas 3 y 4) para N° y Material
    ws.merge_cells("A3:A4")
    ws["A3"] = "N°"
    
    ws.merge_cells("B3:B4")
    ws["B3"] = "MATERIAL"

    # 2. Combinar celdas horizontales (Columnas C a F) para el título ESTADO
    ws.merge_cells("C3:F3")
    ws["C3"] = "CONDICIÓN / ESTADO FÍSICO"

    # 3. Colocar las sub-cabeceras exactamente en la fila 4
    ws["C4"] = "BUEN ESTADO"
    ws["D4"] = "REGULAR"
    ws["E4"] = "MAL ESTADO"
    ws["F4"] = "SE REQUIERE"

    # 4. Pintar y dar estilo a TODA la zona de cabeceras (Filas 3 y 4, Columnas A hasta F)
    for row_idx in range(3, 5):
        for col_idx in range(1, 7):
            celda = ws.cell(row=row_idx, column=col_idx)
            celda.fill = fill_cabecera
            celda.font = font_cabecera
            celda.alignment = align_centro
            celda.border = borde

    # ==========================================
    # IMPRIMIR DATOS 
    # ==========================================
    # 💥 IMPORTANTE: Como la cabecera ahora ocupa las filas 3 y 4, los datos empiezan en la fila 5
    fila_actual = 5 
    
    for indice, fila in enumerate(filas_datos, start=1):
        # Como quitamos la "categoría", ahora los índices van seguidos del 0 al 4
        # fila[0]=Nombre, fila[1]=Bueno, fila[2]=Regular, fila[3]=Malo, fila[4]=Requiere
        fila_con_numero = (indice, fila[0], fila[1], fila[2], fila[3], fila[4]) 
        
        for col_idx, valor in enumerate(fila_con_numero, start=1):
            celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
            celda.font = font_cuerpo
            celda.border = borde
            # Alinear a la izquierda solo el nombre del material (col_idx == 2), el resto centrado
            celda.alignment = align_izq if col_idx == 2 else align_centro
            
        fila_actual += 1

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8  # Ancho pequeño para N°
    ws.column_dimensions['B'].width = 40 # Ancho grande para Nombre
    for l in ['C', 'D', 'E', 'F']: ws.column_dimensions[l].width = 15

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Inventario_{"Aula" if aula_id else "Global"}.xlsx'
    wb.save(response)
    return response

@login_required
@require_POST
def guardar_material_ajax(request):
    """ Función Maestra: Crea o Edita un material dependiendo de si recibe un ID """
    material_id = request.POST.get('id')
    nombre = request.POST.get('nombre')
    
    if not nombre:
        return JsonResponse({'success': False, 'mensaje': 'Debe ingresar el nombre del material.'})
    
    # 1. MODO EDICIÓN
    if material_id:
        material = get_object_or_404(CatalogoMaterial, id=material_id)
        # Verificar que el nuevo nombre no choque con otro existente
        if CatalogoMaterial.objects.filter(nombre__iexact=nombre).exclude(id=material_id).exists():
            return JsonResponse({'success': False, 'mensaje': 'Ya existe otro material con ese nombre.'})
        
        material.nombre = nombre.upper()
        material.save()
        return JsonResponse({'success': True, 'mensaje': 'Material actualizado con éxito.'})
        
    # 2. MODO CREACIÓN
    else:
        material, creado = CatalogoMaterial.objects.get_or_create(
            nombre__iexact=nombre, 
            defaults={'nombre': nombre.upper(), 'activo': True}
        )
        if creado:
            return JsonResponse({'success': True, 'mensaje': f'Material {nombre.upper()} agregado con éxito.'})
        return JsonResponse({'success': False, 'mensaje': 'Este material ya existe en el catálogo.'})

@login_required
@require_POST
def eliminar_material_ajax(request):
    """ Elimina un material validando que no tenga cantidades físicas registradas """
    material_id = request.POST.get('id')
    material = get_object_or_404(CatalogoMaterial, id=material_id)
    
    # 1. Buscamos si hay algún inventario de este material que tenga números mayores a 0
    inventarios_con_datos = material.inventarios_aula.filter(
        Q(buen_estado__gt=0) | 
        Q(regular__gt=0) | 
        Q(mal_estado__gt=0) | 
        Q(se_requiere__gt=0)
    )
    
    # 2. Si existe al menos un aula con cantidades reales, bloqueamos la eliminación
    if inventarios_con_datos.exists():
        return JsonResponse({
            'success': False, 
            'mensaje': 'No se puede eliminar: hay aulas que ya reportaron cantidades físicas de este material.'
        })
    
    # 3. Si pasamos el filtro, significa que todos los registros están en 0 (o no hay registros).
    # Como el modelo tiene on_delete=models.PROTECT, debemos limpiar manualmente los registros en 0 primero.
    material.inventarios_aula.all().delete()
    
    # 4. Finalmente eliminamos el material del catálogo maestro
    material.delete()
    
    return JsonResponse({'success': True, 'mensaje': 'Material eliminado permanentemente.'})

def obtener_textos_actitudinales(notas_dict):
    """ Función Helper que devuelve el texto exacto según el rango de la nota """
    textos_finales = {}
    
    # Textos predefinidos (Rúbrica)
    rubrica = {
        'Puntualidad': {
            'excelente': 'Se presenta a diario y puntualmente a cada clase cumpliendo el horario establecido.',
            'bueno': 'Llega a tiempo a la mayoría de sus clases, mostrando un buen hábito de puntualidad.',
            'adecuado': 'Presenta algunas demoras ocasionales al ingresar a las sesiones de clase.',
            'regular': 'Frecuentemente ingresa tarde a clases, requiere mayor organización de sus tiempos.',
            'proceso': 'Presenta constantes tardanzas que afectan su desarrollo continuo en las sesiones.'
        },
        'Presentacion': {
            'excelente': 'Asiste correctamente uniformado y cumple impecablemente con el aseo personal.',
            'bueno': 'Mantiene una buena presentación personal y uso adecuado del uniforme escolar.',
            'adecuado': 'Cumple con el uniforme de forma básica, con ligeros detalles a mejorar en el aseo.',
            'regular': 'A menudo olvida portar correctamente el uniforme o cuidar su presentación personal.',
            'proceso': 'Requiere supervisión constante de sus padres para cumplir con el aseo y el uso del uniforme.'
        },
        'Participacion': {
            'excelente': 'Se involucra activamente elaborando interrogantes y aportes valiosos en clase.',
            'bueno': 'Participa frecuentemente respondiendo a las preguntas que se le formulan.',
            'adecuado': 'Su participación es esporádica, interactuando principalmente cuando se le solicita.',
            'regular': 'Muestra poca iniciativa para participar o integrarse en las dinámicas del aula.',
            'proceso': 'No participa en las sesiones, requiriendo estímulo constante para integrarse.'
        },
        'Disciplina': {
            'excelente': 'Está plenamente comprometido respetando las normas y sugerencias del Coach.',
            'bueno': 'Muestra un buen comportamiento general y respeta a sus compañeros y docentes.',
            'adecuado': 'Su conducta es aceptable, aunque requiere algunos recordatorios sobre las normas.',
            'regular': 'Le cuesta mantener la concentración y a veces interrumpe el desarrollo de la clase.',
            'proceso': 'Incumple frecuentemente los acuerdos de convivencia, afectando el clima del aula.'
        },
        'Responsabilidad': {
            'excelente': 'Otorga en el plazo indicado íntegramente los trabajos solicitados. Tiene todo al día.',
            'bueno': 'Cumple con la entrega de sus deberes en los plazos establecidos la mayoría de las veces.',
            'adecuado': 'Entrega sus tareas, pero en ocasiones presenta retrasos o trabajos incompletos.',
            'regular': 'Suele olvidar sus materiales, cuadernos o la entrega de trabajos asignados.',
            'proceso': 'No cumple con las actividades académicas asignadas, afectando su aprendizaje.'
        }
    }

    # Lógica para elegir el rango
    for criterio, nota in notas_dict.items():
        if nota >= 19: nivel = 'excelente'     # 91-100%
        elif nota >= 17: nivel = 'bueno'       # 81-90%
        elif nota >= 14: nivel = 'adecuado'    # 66-80%
        elif nota >= 11: nivel = 'regular'     # 51-65%
        else: nivel = 'proceso'                # 0-50%
        
        # En caso de que el tutor no haya puesto notas (0), dejamos el mensaje genérico
        if nota == 0:
            textos_finales[criterio] = "Aún no se han registrado evaluaciones para este criterio en el presente bimestre."
        else:
            textos_finales[criterio] = rubrica[criterio][nivel]

    return textos_finales

def reporte_progresivo_pdf(request, matricula_id):
    """ Genera el Informe Técnico Pedagógico Progresivo con Datos Reales """
    
    # 1. Obtenemos la matrícula exacta
    matricula = get_object_or_404(Matricula, id=matricula_id)
    estudiante = matricula.estudiante
    aula = matricula.aula
    
    # 2. 💥 LÓGICA DINÁMICA: Obtenemos el bimestre del periodo activo
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    # Si no hay periodo activo por alguna razón, usamos 'I' por seguridad
    bimestre_activo = periodo_actual.bimestre_actual if periodo_actual else 'I'
    
    # 3. DATOS REALES: Buscamos las notas actitudinales del alumno
    try:
        
        evaluacion = EvaluacionActitudinal.objects.get(
            matricula=matricula, 
            bimestre=bimestre_activo
        )
        notas_actitudinales = {
            'Puntualidad': evaluacion.puntualidad,
            'Presentacion': evaluacion.presentacion,   # Sin tilde
            'Participacion': evaluacion.cuidado_patrimonio, # Sin tilde
            'Disciplina': evaluacion.orden_limpieza,
            'Responsabilidad': evaluacion.respeto_normas
        }
        
        # Leemos si la IA ya hizo el trabajo antes
        recomendaciones_guardadas = []
        if evaluacion.recomendacion_ia:
            recomendaciones_guardadas = json.loads(evaluacion.recomendacion_ia)
            
    except EvaluacionActitudinal.DoesNotExist:
        # Si el tutor aún no registra notas, enviamos todo en 0
        evaluacion = None
        notas_actitudinales = {
            'Puntualidad': 0, 'Presentación': 0, 'Participación': 0, 'Disciplina': 0, 'Responsabilidad': 0
        }
        recomendaciones_guardadas = []

    textos_tabla = obtener_textos_actitudinales(notas_actitudinales)

    context = {
        'matricula': matricula,
        'estudiante': estudiante,
        'aula': aula,
        'notas_actitudinales': notas_actitudinales,
        'textos_tabla': textos_tabla,
        'evaluacion_id': evaluacion.id if evaluacion else None, # ID para que el JS sepa a quién llamar
        'recomendaciones_ia': recomendaciones_guardadas, # Puede estar vacío
        'bimestre_actual': bimestre_activo,
    }
    
    return render(request, 'academico/reporte_progresivo.html', context)

# 2. LA VISTA ASÍNCRONA (AJAX)
@csrf_exempt # Evita errores de seguridad al llamar por AJAX
def generar_recomendacion_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            evaluacion_id = data.get('evaluacion_id')
            evaluacion = EvaluacionActitudinal.objects.get(id=evaluacion_id)
            
            notas_actitudinales = {
                'Puntualidad': evaluacion.puntualidad, 'Presentacion': evaluacion.presentacion,
                'Participacion': evaluacion.cuidado_patrimonio, 'Disciplina': evaluacion.orden_limpieza,
                'Responsabilidad': evaluacion.respeto_normas
            }
            nombre_alumno = evaluacion.matricula.estudiante.nombres
            
            # Llamamos a Gemini
            recomendaciones = generar_4_recomendaciones_ia(nombre_alumno, notas_actitudinales)
            
            # 💥 GUARDAMOS en la BD para que NUNCA MÁS se vuelva a generar
            evaluacion.recomendacion_ia = json.dumps(recomendaciones)
            evaluacion.save()
            
            return JsonResponse({'status': 'success', 'recomendaciones': recomendaciones})
        except Exception as e:
            return JsonResponse({'status': 'error', 'mensaje': str(e)})
    return JsonResponse({'status': 'error'})

def exportar_simulacro_word(request, simulacro_id):
    """ Motor Generador del Simulacro en formato Word (.docx) """
    simulacro = get_object_or_404(Simulacro, id=simulacro_id)
    
    # 1. Cargamos la plantilla base desde tu carpeta
    ruta_plantilla = os.path.join(settings.BASE_DIR, 'templates_archivos', 'plantilla_simulacro.docx')
    doc = DocxTemplate(ruta_plantilla)
    
    # Intentamos obtener al tutor buscando el aula correspondiente a ese grado y nivel
    aula_ref = Aula.objects.filter(grado=simulacro.grado, nivel=simulacro.nivel).first()
    # ✅ LÍNEA CORREGIDA:
    nombre_tutor = f"{aula_ref.tutor.nombres} {aula_ref.tutor.apellidos}" if (aula_ref and aula_ref.tutor) else "________________________"

    # 2. Obtenemos las preguntas ordenadas por área y nombre del curso
    preguntas_qs = simulacro.preguntas.select_related('curso').order_by('curso__area', 'curso__nombre', 'id')
    
    # 3. Estructuramos los datos para inyectarlos (Área -> Curso -> Preguntas)
    areas_dict = {}
    
    for i, preg in enumerate(preguntas_qs, start=1):
        area_code = preg.curso.area
        area_name = preg.curso.get_area_display().upper() # Ej: APTITUD ACADÉMICA
        curso_name = preg.curso.nombre.upper()            # Ej: RAZONAMIENTO MATEMÁTICO
        
        if area_code not in areas_dict:
            areas_dict[area_code] = {'nombre': area_name, 'cursos_map': {}}
            
        if curso_name not in areas_dict[area_code]['cursos_map']:
            areas_dict[area_code]['cursos_map'][curso_name] = {'nombre': curso_name, 'preguntas': []}
            
        # 4. Procesamiento de la Imagen (¡Vital!)
        imagen_inline = None
        if preg.imagen and preg.imagen.name: # Verificamos que sí haya subido un archivo
            ruta_imagen = preg.imagen.path
            if os.path.exists(ruta_imagen):
                # InlineImage incrusta la foto. Mm(70) limita su ancho a 7 centímetros para no desbordar la columna
                imagen_inline = InlineImage(doc, ruta_imagen, width=Mm(70))
        
        # Función interna para pintar de amarillo si es la clave correcta
        def formatear_opcion(texto_opcion, letra_opcion):
            rt = RichText()
            if preg.respuesta_correcta == letra_opcion:
                rt.add(texto_opcion, highlight='yellow')
            else:
                rt.add(texto_opcion)
            return rt

        # 5. Empaquetamos la pregunta con el resaltador automático
        preg_data = {
            'numero': i, 
            'enunciado': preg.enunciado,
            'tiene_imagen': bool(imagen_inline),
            'imagen': imagen_inline,
            'opcion_a': formatear_opcion(preg.opcion_a, 'a'),
            'opcion_b': formatear_opcion(preg.opcion_b, 'b'),
            # Solo mandamos texto si el profesor llenó la opción
            'opcion_c': formatear_opcion(preg.opcion_c, 'c') if preg.opcion_c else None,
            'opcion_d': formatear_opcion(preg.opcion_d, 'd') if preg.opcion_d else None,
            'opcion_e': formatear_opcion(preg.opcion_e, 'e') if preg.opcion_e else None,
        }
        
        areas_dict[area_code]['cursos_map'][curso_name]['preguntas'].append(preg_data)
        
    # 6. Convertimos el diccionario a la lista plana que espera la plantilla Word
    areas_list = []
    for a_code, a_data in areas_dict.items():
        cursos_list = [c_data for c_name, c_data in a_data['cursos_map'].items()]
        areas_list.append({
            'nombre': a_data['nombre'],
            'cursos': cursos_list
        })
        
    # 7. Preparamos el contexto final y renderizamos
    context = {
        'titulo_simulacro': simulacro.titulo.upper(), # Ej: II CONCURSO DE APTITUD ACADÉMICA
        'subtitulo_simulacro': f"{simulacro.periodo.anio} – {simulacro.grado.upper()}", # Ej: 2026 - 6TO GRADO
        'tutor_nombre': nombre_tutor,
        'areas': areas_list
    }
    doc.render(context)
    
    # 8. Forzamos la descarga del archivo generado
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename="Simulacro_{simulacro.grado}_{simulacro.mes}.docx"'
    doc.save(response)
    
    return response

def lista_simulacros(request):
    """ Panel de control de simulacros para la Coordinadora """
    periodo_activo = PeriodoLectivo.objects.filter(activo=True).first()
    # Traemos los simulacros del año escolar actual
    simulacros = Simulacro.objects.filter(periodo=periodo_activo).order_by('-fecha_examen')
    form = SimulacroForm()
    
    return render(request, 'academico/lista_simulacros.html', {
        'simulacros': simulacros,
        'form': form
    })

def guardar_simulacro_ajax(request):
    """ Guarda o edita un simulacro mediante AJAX """
    if request.method == 'POST':
        simulacro_id = request.POST.get('simulacro_id')
        periodo_activo = PeriodoLectivo.objects.filter(activo=True).first()
        
        if simulacro_id:
            simulacro = get_object_or_404(Simulacro, id=simulacro_id)
            form = SimulacroForm(request.POST, instance=simulacro)
        else:
            simulacro = Simulacro(periodo=periodo_activo)
            form = SimulacroForm(request.POST, instance=simulacro)
            
        if form.is_valid():
            form.save()
            return JsonResponse({'status': 'ok', 'message': 'Simulacro guardado correctamente.'})
        return JsonResponse({'status': 'error', 'errors': form.errors})
        
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'})

def datos_simulacro_ajax(request, pk):
    """ Devuelve los datos de un simulacro para cargarlos en el modal de edición """
    simulacro = get_object_or_404(Simulacro, pk=pk)
    return JsonResponse({
        'id': simulacro.id,
        'titulo': simulacro.titulo,
        'mes': simulacro.mes,
        'grado': simulacro.grado,
        'nivel': simulacro.nivel,
        'fecha_examen': simulacro.fecha_examen.strftime('%Y-%m-%d'),
        # 💥 Agregamos esta línea
        'preguntas_esperadas': simulacro.preguntas_esperadas,
        'activo': simulacro.activo
    })

@login_required
def mis_simulacros_docente(request):
    """ Muestra al profesor los simulacros institucionales que están activos """
    docente = get_object_or_404(Personal, user=request.user, cargo='DOC')
    periodo_activo = PeriodoLectivo.objects.filter(activo=True).first()
    
    # Buscamos simulacros del año escolar actual que estén abiertos para recibir preguntas
    simulacros_activos = Simulacro.objects.filter(periodo=periodo_activo, activo=True).order_by('fecha_examen')
    
    return render(request, 'personal/mis_simulacros.html', {
        'simulacros': simulacros_activos,
        'docente': docente
    })

@login_required
def cargar_preguntas_simulacro(request, simulacro_id):
    docente = get_object_or_404(Personal, user=request.user, cargo='DOC')
    simulacro = get_object_or_404(Simulacro, id=simulacro_id)
    
    # 💥 1. Obtenemos los IDs de los cursos que el docente ya finalizó/bloqueó
    cursos_finalizados_ids = list(EntregaSimulacro.objects.filter(
        simulacro=simulacro, docente=docente, finalizado=True
    ).values_list('curso_id', flat=True))
    
    # 2. Obtenemos todas sus asignaciones
    asignaciones = AsignacionAcademica.objects.filter(
        personal=docente,
        aula__grado=simulacro.grado,
        aula__nivel=simulacro.nivel,
        periodo=simulacro.periodo
    ).select_related('curso')
    
    # 💥 3. Filtramos: Solo dejamos los cursos que NO están en la lista de finalizados
    cursos_disponibles = list(set([
        asig.curso for asig in asignaciones if asig.curso.id not in cursos_finalizados_ids
    ]))
    
    # 💥 (Línea nueva) Obtenemos los objetos completos de los cursos finalizados
    cursos_finalizados = list(set([
        asig.curso for asig in asignaciones if asig.curso.id in cursos_finalizados_ids
    ]))

    if request.method == 'POST':
        pregunta_id = request.POST.get('pregunta_id')
        if pregunta_id:
            pregunta = get_object_or_404(PreguntaSimulacro, id=pregunta_id)
            form = PreguntaSimulacroForm(request.POST, request.FILES, instance=pregunta)
        else:
            form = PreguntaSimulacroForm(request.POST, request.FILES)
            
        if form.is_valid():
            nueva_pregunta = form.save(commit=False)
            nueva_pregunta.simulacro = simulacro
            nueva_pregunta.docente = docente
            nueva_pregunta.save()
            return redirect('academico:cargar_preguntas_simulacro', simulacro_id=simulacro.id)
    else:
        form = PreguntaSimulacroForm()
        # El select del formulario ahora solo mostrará los cursos pendientes
        form.fields['curso'].queryset = Curso.objects.filter(id__in=[c.id for c in cursos_disponibles])

    preguntas_guardadas = PreguntaSimulacro.objects.filter(simulacro=simulacro, docente=docente).order_by('curso__nombre', 'id')

    return render(request, 'personal/cargar_preguntas.html', {
        'simulacro': simulacro,
        'form': form,
        'preguntas': preguntas_guardadas,
        'cursos_disponibles': cursos_disponibles,
        'cursos_finalizados_ids': cursos_finalizados_ids,
        'cursos_finalizados': cursos_finalizados, # 💥 Enviamos la lista de objetos
    })

@login_required
def datos_pregunta_ajax(request, pk):
    pregunta = get_object_or_404(PreguntaSimulacro, pk=pk)
    return JsonResponse({
        'id': pregunta.id,
        'curso_id': pregunta.curso.id,
        'enunciado': pregunta.enunciado,
        'opcion_a': pregunta.opcion_a,
        'opcion_b': pregunta.opcion_b,
        'opcion_c': pregunta.opcion_c or '',
        'opcion_d': pregunta.opcion_d or '',
        'opcion_e': pregunta.opcion_e or '',
        'respuesta_correcta': pregunta.respuesta_correcta,
        'imagen_url': pregunta.imagen.url if pregunta.imagen else None
    })

@login_required
@require_POST
def eliminar_pregunta_simulacro(request, pregunta_id):
    """
    Elimina una pregunta de un simulacro vía AJAX y devuelve un JSON.
    """
    try:
        # 1. Buscamos la pregunta de forma segura
        pregunta = get_object_or_404(PreguntaSimulacro, id=pregunta_id)
        
        # 2. Eliminamos el registro de la base de datos
        pregunta.delete()
        
        # 3. Respondemos a tu función global JS con el formato que espera
        return JsonResponse({
            'success': True,
            'status': 'ok',
            'message': 'La pregunta fue eliminada con éxito.'
        })
        
    except Exception as e:
        # Si algo falla (ej. base de datos caída), evitamos que el servidor explote
        return JsonResponse({
            'success': False,
            'status': 'error',
            'message': 'Hubo un problema al intentar eliminar la pregunta.'
        }, status=500)

# 1. ACCIÓN DEL DOCENTE: MARCAR COMO COMPLETADO (AJAX)
@csrf_exempt
@login_required
def finalizar_envio_curso_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            simulacro_id = data.get('simulacro_id')
            curso_id = data.get('curso_id')
            docente = get_object_or_404(Personal, user=request.user, cargo='DOC')
            
            # Buscamos o creamos el token de entrega
            entrega, created = EntregaSimulacro.objects.get_or_create(
                simulacro_id=simulacro_id,
                curso_id=curso_id,
                docente=docente
            )
            entrega.finalizado = True
            entrega.save()
            return JsonResponse({'status': 'ok', 'message': 'Curso marcado como completado.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error'})

# 2. PANEL DE LA COORDINADORA: VER QUIÉN FALTA
@login_required
def monitoreo_simulacro(request, simulacro_id):
    """ Muestra una lista de todos los cursos que DEBERÍAN subir preguntas y su estado """
    simulacro = get_object_or_404(Simulacro, id=simulacro_id)
    
    asignaciones = AsignacionAcademica.objects.filter(
        aula__grado=simulacro.grado,
        aula__nivel=simulacro.nivel,
        periodo=simulacro.periodo
    ).select_related('curso', 'personal').distinct()
    
    # 💥 OPTIMIZACIÓN SENIOR: Traemos todos los datos en 2 consultas masivas
    # 1. Contamos las preguntas de todos los profesores en este simulacro a la vez
    preguntas_db = PreguntaSimulacro.objects.filter(simulacro=simulacro).values('curso_id', 'docente_id')
    mapa_preguntas = {}
    for p in preguntas_db:
        llave = (p['curso_id'], p['docente_id'])
        mapa_preguntas[llave] = mapa_preguntas.get(llave, 0) + 1
        
    # 2. Traemos quiénes ya entregaron en un "Set" rápido
    entregas_db = set(EntregaSimulacro.objects.filter(
        simulacro=simulacro, finalizado=True
    ).values_list('curso_id', 'docente_id'))
    
    lista_monitoreo = []
    
    for asig in asignaciones:
        llave_actual = (asig.curso.id, asig.personal.id)
        
        # 💥 Buscamos en los diccionarios (Tiempo: 0.0001 segundos)
        cant_preguntas = mapa_preguntas.get(llave_actual, 0)
        estado_entrega = llave_actual in entregas_db
        
        lista_monitoreo.append({
            'curso': asig.curso.nombre,
            'docente': f"{asig.personal.apellidos}, {asig.personal.nombres}",
            'cantidad': cant_preguntas,
            'finalizado': estado_entrega
        })
        
    return render(request, 'academico/monitoreo_simulacro.html', {
        'simulacro': simulacro,
        'reporte': lista_monitoreo
    })

@csrf_exempt
@login_required
def reabrir_envio_curso_ajax(request):
    """ Permite al docente desbloquear un curso si se equivocó, siempre que el examen siga abierto """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            simulacro_id = data.get('simulacro_id')
            curso_id = data.get('curso_id')
            docente = get_object_or_404(Personal, user=request.user, cargo='DOC')
            
            # Validamos que la coordinadora no haya cerrado todo el examen
            simulacro = get_object_or_404(Simulacro, id=simulacro_id)
            if not simulacro.activo:
                return JsonResponse({'status': 'error', 'message': 'El simulacro ya fue cerrado por Coordinación. No puedes hacer más cambios.'})

            # Eliminamos el registro de "Entrega" para que vuelva a estar pendiente
            EntregaSimulacro.objects.filter(
                simulacro_id=simulacro_id,
                curso_id=curso_id,
                docente=docente
            ).delete()
            
            return JsonResponse({'status': 'ok', 'message': 'Curso reabierto para edición.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error'})

def configuracion_institucion(request):
    institucion, created = Institucion.objects.get_or_create(pk=1)
    
    if request.method == 'POST':
        # 💥 Si el usuario hizo clic en la 'X', eliminamos el logo de la BD
        if request.POST.get('clear_logo') == 'true':
            if institucion.logo:
                institucion.logo.delete()
                
        form = InstitucionForm(request.POST, request.FILES, instance=institucion)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuración actualizada con éxito.")
            return redirect('academico:configuracion')
    else:
        form = InstitucionForm(instance=institucion)
        
    return render(request, 'academico/configuracion.html', {
        'form': form, 
        'institucion': institucion
    })