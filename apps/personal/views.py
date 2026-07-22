from django.contrib import messages
from django import forms

from django.shortcuts import redirect, render, get_object_or_404
import gspread
from google.oauth2.service_account import Credentials
from apps.academico.models import ArchivoMaterial, AsignacionAcademica, EntregaSimulacro, Evaluacion, Matricula, Nota, PeriodoLectivo, SolicitudImpresion, Aula, EvaluacionActitudinal, CierreRegistroBimestral
from apps.asistencia.models import AsistenciaEstudiante
from .models import Personal
from .forms import EditarPerfilForm, PersonalForm
from django.http import JsonResponse
import json
from django.views.decorators.http import require_POST
from django.urls import reverse
from collections import Counter
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
import os
from django.conf import settings
from openpyxl.drawing.image import Image as OpenpyxlImage
from django.contrib.auth.views import LoginView


def obtener_personal_logueado(request):
    """Obtiene el personal logueado y verifica que siga activo en la institución."""
    # 💥 Añadimos el filtro estado='Activo' directamente a la consulta SQL
    personal_actual = Personal.objects.filter(user=request.user, estado='Activo').first()
    if not personal_actual:
        return None
    return personal_actual

def raiz_redireccion(request):
    """Controla el acceso a la URL raíz del servidor."""
    if request.user.is_authenticated:
        return redirect('personal:enrutador_principal')
    return redirect('login')

def lista_personal(request):
    personal = Personal.objects.all()
    form = PersonalForm()
    return render(request, 'academico/lista_personal.html', {'personal': personal, 'form': form})

class LoginPersonalizadoView(LoginView):
    """
    Vista maestra de autenticación. 
    Intercepta el formulario para configurar la duración de la sesión.
    """
    template_name = 'registration/login.html'
    redirect_authenticated_user = True

    def form_valid(self, form):
        # 1. Hacemos que Django loguee al usuario normalmente
        response = super().form_valid(form)
        
        # 2. Leemos si el switch "Mantener sesión iniciada" llegó marcado
        remember_me = self.request.POST.get('remember_me', None)

        if remember_me == 'on':
            # Si marcó la casilla: La sesión durará 2 semanas (1209600 segundos)
            # Esto sobrevive aunque apague la PC o el celular
            self.request.session.set_expiry(1209600)
        else:
            # 💥 Si NO la marcó: La sesión se destruye automáticamente al cerrar la pestaña/navegador
            self.request.session.set_expiry(0)

        return response

def guardar_personal_ajax(request):
    data = {}
    if request.method == 'POST':
        pk = request.POST.get('personal_id')
        if pk: # Si hay ID, es EDITAR
            personal = Personal.objects.get(pk=pk)
            form = PersonalForm(request.POST, instance=personal)
            data['message'] = "¡Personal actualizado correctamente!"
        else: # Si no hay ID, es CREAR
            form = PersonalForm(request.POST)
            data['message'] = "¡Personal guardado correctamente!"

        if form.is_valid():
            form.save()
            data['status'] = 'ok'
        else:
            data['status'] = 'error'
            data['errors'] = form.errors
    return JsonResponse(data)

def obtener_personal_data(request, pk):
    personal = Personal.objects.get(pk=pk)
    data = {
        'id': personal.id,
        'dni': personal.dni,
        'nombres': personal.nombres,
        'apellidos': personal.apellidos,
        'tipo_contrato': personal.tipo_contrato,
        'cargo': personal.cargo,
        'telefono': personal.telefono,
        'correo': personal.correo,
        'fecha_ingreso': personal.fecha_ingreso,
    }
    return JsonResponse(data)

@require_POST
def eliminar_personal_ajax(request, pk):
    personal = get_object_or_404(Personal, pk=pk)
    personal.delete()
    return JsonResponse({'status': 'ok', 'message': 'Personal eliminado correctamente.'})

@require_POST
def cambiar_estado_personal_ajax(request, pk):
    personal = get_object_or_404(Personal, pk=pk)
    nuevo_estado = request.POST.get('nuevo_estado')
    if nuevo_estado in dict(Personal.ESTADOS):
        personal.estado = nuevo_estado
        personal.save()
        return JsonResponse({'status': 'ok', 'message': 'Estado actualizado correctamente.'})
    return JsonResponse({'status': 'error', 'message': 'Estado no válido.'})

@login_required
@never_cache
def mis_cursos(request):
    # 1. Identificamos quién es el usuario logueado
    personal_actual = obtener_personal_logueado(request)
    
    if not personal_actual:
        return render(request, 'errores/sin_perfil.html', {
            'mensaje': 'Su cuenta de usuario actual no tiene asignado un perfil en la tabla de Personal.'
        })
    
    # 2. 💥 EL FILTRO DE SEGURIDAD MODULAR: Validamos si realmente es un Docente
    if personal_actual.cargo not in ['DOC', 'Docente']:
        return render(request, 'errores/sin_perfil.html', {
            'mensaje': 'Acceso Restringido. El módulo de "Mis Cursos" es de uso exclusivo para el Personal Docente.'
        })
    
    # En tu modelo AsignacionAcademica, el campo 'personal' ahora apunta a 'Personal'
    asignaciones = AsignacionAcademica.objects.filter(
        personal=personal_actual,
        periodo__activo=True
    ).select_related('curso', 'aula').order_by('curso__nombre', 'aula__nivel', 'aula__grado')

    # Mantenemos 'personal/...' si decidiste renombrar la carpeta de templates
    return render(request, 'personal/mis_cursos.html', {
        'personal': personal_actual,
        'asignaciones': asignaciones
    })

def lista_evaluaciones(request, asignacion_id):
    # Obtenemos el curso y aula que el profe seleccionó
    asignacion = get_object_or_404(AsignacionAcademica, id=asignacion_id)
    
    # Traemos todas las evaluaciones creadas para este curso, ordenadas de las más recientes a las más antiguas
    evaluaciones = Evaluacion.objects.filter(asignacion=asignacion).order_by('-fecha', '-id')
    
    # 💥 Obtenemos los candados de este curso para saber qué bimestres están cerrados
    cierres_db = CierreRegistroBimestral.objects.filter(asignacion=asignacion)
    # Convertimos a un diccionario { 'I': True, 'II': False } para que el HTML lo lea fácil
    diccionario_cierres = {c.bimestre: c.cerrado for c in cierres_db}
    
    # 💥 LA MAGIA: Preparamos la data masticada para el HTML
    bimestres_info = []
    for b_key, b_name in asignacion.periodo.BIMESTRES:
        bimestres_info.append({
            'key': b_key,
            'name': b_name,
            'cerrado': diccionario_cierres.get(b_key, False)
        })
    
    return render(request, 'personal/lista_evaluaciones.html', {
        'asignacion': asignacion,
        'evaluaciones': evaluaciones,
        'diccionario_cierres': diccionario_cierres, # Lo mandamos al frontend
        'bimestres_info': bimestres_info # Lo mandamos al frontend
    })

@require_POST
def guardar_evaluacion_ajax(request):
    asignacion_id = request.POST.get('asignacion_id')
    tipo = request.POST.get('tipo')
    
    asignacion = get_object_or_404(AsignacionAcademica, id=asignacion_id)
    
    # 1. SACAMOS EL BIMESTRE DIRECTO DEL AÑO ESCOLAR ACTIVO (El profe no lo toca)
    bimestre_actual = asignacion.periodo.bimestre_actual
    
    # 2. Contamos cuántas evaluaciones existen de ese tipo EN ESE BIMESTRE
    conteo = Evaluacion.objects.filter(
        asignacion=asignacion, 
        tipo=tipo, 
        bimestre=bimestre_actual
    ).count()
    
    numero = conteo + 1 # Si no hay, es el 1. Si hay 2, es el 3.
    
    # 3. Generamos el nombre exacto como querías
    if tipo == 'DESAFIO':
        nombre_auto = f"Desafío - Tema {numero}"
    elif tipo == 'SIMULACRO':
        nombre_auto = f"Concurso de Aptitud Mensual {numero}"
    elif tipo == 'MENSUAL':
        nombre_auto = "Control de Calidad" # Sin número porque es 1 por bimestre
    elif tipo == 'BIMESTRAL':
        nombre_auto = "ISO Ingeniería"     # Sin número porque es 1 por bimestre
    elif tipo == 'CUADERNO':
        nombre_auto = f"Cuaderno - Revisión {numero}"
    elif tipo == 'LIBRO':
        nombre_auto = f"Libro - Revisión {numero}"
    else:
        nombre_auto = f"{tipo} {numero}"

    # 4. Creamos la evaluación
    nueva_eval = Evaluacion.objects.create(
        asignacion=asignacion,
        tipo=tipo,
        bimestre=bimestre_actual,
        nombre=nombre_auto
    )

    # 5. Generamos las casillas vacías de notas para los alumnos de esa aula
    estudiantes_matriculados = Matricula.objects.filter(
        aula=asignacion.aula, 
        periodo=asignacion.periodo
    )

    for matricula in estudiantes_matriculados:
        Nota.objects.create(
            matricula=matricula,
            evaluacion=nueva_eval,
            valor=None 
        )

    return JsonResponse({
        'status': 'ok', 
        'message': f'{nombre_auto} creado correctamente.',
        'evaluacion_id': nueva_eval.id
    })
    
def registro_notas(request, evaluacion_id):
    # Traemos la evaluación específica
    evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id)
    
    # Verificamos si este bimestre está cerrado para bloquear los inputs
    cierre = CierreRegistroBimestral.objects.filter(asignacion=evaluacion.asignacion, bimestre=evaluacion.bimestre).first()
    registro_cerrado = cierre.cerrado if cierre else False
    
    # =========================================================
    # LÓGICA DE AUTO-REPARACIÓN (Self-Healing)
    # Evita el error de "alumnos nuevos" o "notas no generadas"
    # =========================================================
    matriculas_actuales = Matricula.objects.filter(
        aula=evaluacion.asignacion.aula, 
        periodo=evaluacion.asignacion.periodo
    )
    
    for matricula in matriculas_actuales:
        # get_or_create busca si ya existe la nota. Si NO existe, la crea en blanco.
        # Así aseguramos que nadie se quede fuera de la lista.
        Nota.objects.get_or_create(
            matricula=matricula,
            evaluacion=evaluacion,
            defaults={'valor': None}
        )
    # =========================================================

    # Ahora sí, traemos TODAS las notas de esta evaluación ordenadas
    notas = Nota.objects.filter(evaluacion=evaluacion).select_related(
        'matricula__estudiante'
    ).order_by('matricula__estudiante__apellidos')
    
    return render(request, 'personal/registro_notas.html', {
        'evaluacion': evaluacion,
        'notas': notas,
        'registro_cerrado': registro_cerrado
    })
    
@require_POST
def guardar_nota_ajax(request):
    nota_id = request.POST.get('nota_id')
    valor = request.POST.get('valor')
    
    try:
        nota = get_object_or_404(Nota, id=nota_id)
        
        # 💥 SEGURIDAD ANTI-HACKERS: Verificamos si el candado está cerrado
        cierre = CierreRegistroBimestral.objects.filter(asignacion=nota.evaluacion.asignacion, bimestre=nota.evaluacion.bimestre).first()
        if cierre and cierre.cerrado:
            return JsonResponse({'status': 'error', 'message': 'El registro está CERRADO. No puede modificar notas.'})
        
        # Si el profesor borra la nota (la deja vacía), guardamos NULL
        if valor == '':
            nota.valor = None
        else:
            # Aseguramos que sea un número decimal
            nota.valor = float(valor)
            
        nota.save()
        return JsonResponse({'status': 'ok', 'message': 'Guardado'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_POST
def toggle_cierre_registro_ajax(request):
    asignacion_id = request.POST.get('asignacion_id')
    bimestre = request.POST.get('bimestre')
    accion = request.POST.get('accion') # 'cerrar' o 'abrir'
    
    asignacion = get_object_or_404(AsignacionAcademica, id=asignacion_id)
    
    cierre, created = CierreRegistroBimestral.objects.get_or_create(
        asignacion=asignacion, bimestre=bimestre
    )
    
    if accion == 'cerrar':
        cierre.cerrado = True
        cierre.fecha_cierre = timezone.now()
        mensaje = f"Registro del Bimestre {bimestre} cerrado y enviado a Coordinación."
    else:
        cierre.cerrado = False
        mensaje = f"Registro del Bimestre {bimestre} reabierto. Ya puede editar notas."
        
    cierre.save()
    return JsonResponse({'success': True, 'mensaje': mensaje})

def material_upload(request, asignacion_id):
    asignacion = get_object_or_404(AsignacionAcademica, id=asignacion_id)
    
    if request.method == 'POST':
        fecha = request.POST.get('fecha_requerida')
        notas = request.POST.get('instrucciones')
        
        # 1. Creamos la cabecera
        solicitud = SolicitudImpresion.objects.create(
            asignacion=asignacion,
            fecha_requerida=fecha,
            instrucciones=notas
        )
        
        # 2. Procesamos los múltiples archivos
        # Los archivos vienen en listas gracias al botón +
        tipos = request.POST.getlist('tipo_archivo[]')
        archivos = request.FILES.getlist('archivo[]')
        
        for i in range(len(archivos)):
            ArchivoMaterial.objects.create(
                solicitud=solicitud,
                tipo=tipos[i],
                archivo=archivos[i]
            )
            
        return redirect('personal:mis_cursos') # O a una lista de materiales

    return render(request, 'personal/material_upload.html', {'asignacion': asignacion})

@login_required
@never_cache
def centro_materiales(request):
    # 1. Identificamos quién es el usuario logueado
    personal_actual = obtener_personal_logueado(request)
    
    if not personal_actual:
        return render(request, 'errores/sin_perfil.html', {
            'mensaje': 'Su cuenta de usuario actual no tiene asignado un perfil en la tabla de Personal.'
        })
    
    # 2. 💥 EL FILTRO DE SEGURIDAD MODULAR: Validamos si realmente es un Docente
    if personal_actual.cargo not in ['DOC', 'Docente']:
        return render(request, 'errores/sin_perfil.html', {
            'mensaje': 'Acceso Restringido. El módulo de "Mis Cursos" es de uso exclusivo para el Personal Docente.'
        })

    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    
    if request.method == 'POST':
        asignacion_id = request.POST.get('asignacion_id')
        tema = request.POST.get('tema')
        instrucciones = request.POST.get('instrucciones')
        
        asignacion = get_object_or_404(AsignacionAcademica, id=asignacion_id)
        
        # 1. Definimos las reglas estrictas de seguridad (Extensiones y Tipos MIME)
        EXTENSIONES_PERMITIDAS = ('.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp')
        MIMES_DOCUMENTOS = [
            'application/pdf',
            'application/msword', # .doc
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document' # .docx
        ]
        
        total_secciones = int(request.POST.get('total_secciones', 0))
        
        # 2. PRIMERA PASADA: Solo verificamos (sin guardar nada aún) para no dejar registros "a medias"
        for i in range(total_secciones):
            archivos = request.FILES.getlist(f'archivos_{i}')
            for f in archivos:
                ext_valida = f.name.lower().endswith(EXTENSIONES_PERMITIDAS)
                mime_valido = f.content_type.startswith('image/') or f.content_type in MIMES_DOCUMENTOS
                
                if not (ext_valida and mime_valido):
                    # Si detecta UN solo archivo invasor, rechaza toda la solicitud
                    return JsonResponse({
                        'success': False,
                        'mensaje': f'El archivo "{f.name}" no está permitido. Por favor, suba únicamente Imágenes, PDFs o documentos Word.'
                    })

        # 3. Si todo está limpio, procedemos a crear los registros en la base de datos
        solicitud = SolicitudImpresion.objects.create(
            personal=personal_actual, 
            asignacion=asignacion,
            bimestre=periodo_actual.bimestre_actual,
            tema=tema,
            instrucciones=instrucciones
        )
        
        tipos_enviados = []
        
        for i in range(total_secciones):
            tipo = request.POST.get(f'tipo_{i}')
            archivos = request.FILES.getlist(f'archivos_{i}')
            tipo_display = dict(ArchivoMaterial.TIPOS).get(tipo, tipo)
            
            for f in archivos:
                ArchivoMaterial.objects.create(solicitud=solicitud, tipo=tipo, archivo=f)
                tipos_enviados.append(tipo_display)
        
        contador_tipos = Counter(tipos_enviados)
        materiales_resumen = [{"tipo": t, "cant": c} for t, c in contador_tipos.items()]
        fecha_local = timezone.localtime(solicitud.fecha_subida)

        return JsonResponse({
            'success': True,
            'registro': {
                'id': solicitud.id,
                'curso': solicitud.asignacion.curso.nombre,
                'tema': solicitud.get_tema_display(),
                'fecha': fecha_local.strftime("%d/%m/%Y %H:%M"), 
                'materiales': materiales_resumen,
                'estado': solicitud.get_estado_display(),
                'url_eliminar': reverse('personal:eliminar_solicitud', args=[solicitud.id])
            }
        })

    # Bloque GET
    asignaciones = AsignacionAcademica.objects.filter(personal=personal_actual, periodo=periodo_actual)
    historial = SolicitudImpresion.objects.filter(personal=personal_actual).prefetch_related('archivos').order_by('-fecha_subida')

    return render(request, 'personal/centro_materiales.html', {
        'personal': personal_actual, 
        'asignaciones': asignaciones,
        'historial': historial,
        'bimestre_actual': periodo_actual.bimestre_actual if periodo_actual else 'I',
    })

# NUEVA FUNCIÓN PARA ELIMINAR UN ENVÍO
def eliminar_solicitud(request, solicitud_id):
    if request.method == 'POST':
        solicitud = get_object_or_404(SolicitudImpresion, id=solicitud_id)
        if solicitud.estado == 'PENDIENTE':
            solicitud.delete()
            return JsonResponse({'success': True, 'mensaje': 'Envío eliminado correctamente'})
        else:
            return JsonResponse({'success': False, 'mensaje': 'Ya está en proceso, no se puede eliminar'})
    return JsonResponse({'success': False, 'mensaje': 'Método no permitido'})

@login_required
def enrutador_principal(request):
    try:
        # MÉTODO BLINDADO: Buscamos directamente en la tabla Personal el registro que tenga este usuario
        personal_actual = Personal.objects.filter(user=request.user).first()
        
        # Si no encuentra nada, significa que es un usuario sin perfil
        if not personal_actual:
            return render(request, 'errores/sin_perfil.html', {
                'mensaje': 'La cuenta con la que ingresó no está asignada a ningún registro de Personal.'
            })

        cargo = personal_actual.cargo
        
        # Redirecciones (Aceptamos tanto el código corto 'DOC' como la palabra completa 'Docente')
        if cargo == 'DOC': 
            return redirect('personal:mis_cursos') 
        elif cargo in ['COO', 'DIR', 'Coordinador', 'Coordinadora', 'Director']: 
            return redirect('core:home') 
        elif cargo in ['SEC', 'Secretaria']: 
            return redirect('core:home')
        elif cargo in ['ASI', 'Asistente']: 
            return redirect('core:home')
        else:
            return render(request, 'errores/sin_perfil.html', {
                'mensaje': f'Usted tiene el cargo "{cargo}", pero su módulo aún no está configurado.'
            })
            
    except Exception as e:
        print(f"🔥 ERROR EN EL ENRUTADOR: {e}") 
        return render(request, 'errores/sin_perfil.html', {
            'mensaje': f'Ocurrió un error técnico: {e}'
        })

@login_required
def registro_actitudinal(request, aula_id, bimestre):
    aula = get_object_or_404(Aula, id=aula_id)
    periodo_actual = PeriodoLectivo.objects.get(activo=True)
    
    # 💥 OPTIMIZACIÓN 1: select_related('estudiante') para no consultar el nombre 30 veces
    matriculas = Matricula.objects.filter(
        aula=aula, 
        periodo=periodo_actual,
        estudiante__estado='Activo'
    ).select_related('estudiante').order_by('estudiante__apellidos') 
    
    # 💥 OPTIMIZACIÓN 2: Extraemos todas las evaluaciones de este bimestre de un solo golpe
    evaluaciones_existentes = EvaluacionActitudinal.objects.filter(
        matricula__in=matriculas, bimestre=bimestre
    )
    mapa_evaluaciones = {e.matricula_id: e for e in evaluaciones_existentes}
    
    lista_estudiantes = []
    nuevas_evaluaciones = []

    for mat in matriculas:
        if mat.id in mapa_evaluaciones:
            eval_act = mapa_evaluaciones[mat.id]
        else:
            # Si no existe, la preparamos en memoria (SIN tocar la base de datos aún)
            eval_act = EvaluacionActitudinal(matricula=mat, bimestre=bimestre)
            nuevas_evaluaciones.append(eval_act)
            
        lista_estudiantes.append({
            'matricula': mat,
            'evaluacion_act': eval_act
        })
        
    # 💥 OPTIMIZACIÓN 3: Creamos todas las que faltaban en 1 solo viaje a la BD
    if nuevas_evaluaciones:
        EvaluacionActitudinal.objects.bulk_create(nuevas_evaluaciones)
        
        # Opcional: Recargamos los IDs frescos si la BD los necesita para el renderizado
        evals_frescas = EvaluacionActitudinal.objects.filter(matricula__in=matriculas, bimestre=bimestre)
        mapa_fresco = {e.matricula_id: e for e in evals_frescas}
        for item in lista_estudiantes:
            item['evaluacion_act'] = mapa_fresco[item['matricula'].id]
        
    return render(request, 'personal/registro_actitudinal.html', {
        'aula': aula,
        'bimestre': bimestre,
        'lista_estudiantes': lista_estudiantes
    })

def guardar_actitudinal_ajax(request):
    if request.method == 'POST':
        eval_id = request.POST.get('eval_id')
        campo = request.POST.get('campo')
        valor = request.POST.get('valor')
        
        try:
            evaluacion = EvaluacionActitudinal.objects.get(id=eval_id)
            # Actualizamos el campo dinámicamente
            setattr(evaluacion, campo, int(valor))
            
            # 💥 LA MAGIA: Borramos la recomendación antigua de la IA 
            # para obligar a que Gemini genere una nueva la próxima vez
            evaluacion.recomendacion_ia = None 
            
            evaluacion.save()
            
            return JsonResponse({
                'status': 'ok', 
                'nuevo_promedio': evaluacion.promedio_actitudinal
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

@login_required
def matriz_notas(request, asignacion_id):
    """ Genera la Sábana de Notas dinámica para el profesor """
    asignacion = get_object_or_404(AsignacionAcademica, id=asignacion_id)
    # 1. Primero obtenemos el periodo lectivo que está activo en el colegio
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    
    # 💥 LA SOLUCIÓN: Si hay un periodo activo, extraemos su bimestre actual.
    # Si por alguna razón no hay ninguno activo, usamos 'I' como salvavidas.
    bimestre_predeterminado = periodo_actual.bimestre_actual if periodo_actual else 'I'
    
    # 2. Capturamos el parámetro de la URL. Si viene vacío, adoptará el del sistema (ej: 'II')
    bimestre = request.GET.get('bimestre', bimestre_predeterminado)

    # 1. Traemos las evaluaciones del profesor ordenadas cronológicamente
    evaluaciones = asignacion.evaluaciones.filter(bimestre=bimestre).order_by('fecha', 'id')
    
    # Las agrupamos por tipo para las cabeceras HTML
    evals_cuaderno = evaluaciones.filter(tipo='CUADERNO')
    evals_desafio = evaluaciones.filter(tipo='DESAFIO')
    evals_examenes = evaluaciones.filter(tipo__in=['MENSUAL', 'BIMESTRAL', 'SIMULACRO'])

    # 2. Traemos a todos los alumnos del aula
    # ✅ LÍNEA CORREGIDA
    matriculas = Matricula.objects.filter(aula=asignacion.aula, estudiante__estado='Activo').select_related('estudiante').order_by('estudiante__apellidos')
    
    # 3. Traemos TODAS las notas de este curso en una sola consulta optimizada
    notas_db = Nota.objects.filter(evaluacion__in=evaluaciones).select_related('matricula', 'evaluacion')
    
    # Convertimos las notas en un diccionario rápido: dict[matricula_id][evaluacion_id] = valor
    diccionario_notas = {}
    for n in notas_db:
        if n.matricula_id not in diccionario_notas:
            diccionario_notas[n.matricula_id] = {}
        diccionario_notas[n.matricula_id][n.evaluacion_id] = n.valor

    # 4. Construimos la data cruzada calculando promedios
    datos_matriz = []
    for mat in matriculas:
        notas_alumno = diccionario_notas.get(mat.id, {})

        # Función interna para calcular promedio ignorando los "Null" (exámenes no dados)
        def calcular_promedio(grupo_evaluaciones):
            valores = [notas_alumno[e.id] for e in grupo_evaluaciones if e.id in notas_alumno and notas_alumno[e.id] is not None]
            return round(sum(valores) / len(valores), 2) if valores else 0.0

        prom_cuaderno = calcular_promedio(evals_cuaderno)
        prom_desafio = calcular_promedio(evals_desafio)
        prom_examen = calcular_promedio(evals_examenes)

        # Promedio General (Puedes cambiar esta fórmula si el colegio usa pesos diferentes)
        sumatoria_promedios = [p for p in [prom_cuaderno, prom_desafio, prom_examen] if p > 0]
        prom_general = round(sum(sumatoria_promedios) / len(sumatoria_promedios), 2) if sumatoria_promedios else 0.0

        datos_matriz.append({
            'estudiante': f"{mat.estudiante.apellidos}, {mat.estudiante.nombres}",
            'notas': notas_alumno, # Diccionario de ID_evaluacion -> Nota
            'prom_cuaderno': prom_cuaderno,
            'prom_desafio': prom_desafio,
            'prom_examen': prom_examen,
            'prom_general': prom_general
        })

    return render(request, 'personal/matriz_notas.html', {
        'asignacion': asignacion,
        'bimestre': bimestre,
        'evals_cuaderno': evals_cuaderno,
        'evals_desafio': evals_desafio,
        'evals_examenes': evals_examenes,
        'datos_matriz': datos_matriz,
    })

@login_required
def reporte_agenda_semanal(request, aula_id):
    """ Genera los tickets de agendas filtrando por TEMA """
    aula = get_object_or_404(Aula, id=aula_id)
    asignaciones = AsignacionAcademica.objects.filter(aula=aula)

    temas_disponibles = Evaluacion.objects.filter(
        asignacion__in=asignaciones, tipo='DESAFIO'
    ).values_list('nombre', flat=True).distinct().order_by('nombre')

    tema_seleccionado = request.GET.get('tema', '')
    if not tema_seleccionado and temas_disponibles:
        tema_seleccionado = temas_disponibles[0]

    evaluaciones = Evaluacion.objects.filter(
        asignacion__in=asignaciones,
        tipo='DESAFIO',
        nombre=tema_seleccionado
    )

    matriculas = Matricula.objects.filter(aula=aula, estudiante__estado='Activo').select_related('estudiante').order_by('estudiante__apellidos')
    notas_db = Nota.objects.filter(evaluacion__in=evaluaciones).select_related('matricula', 'evaluacion__asignacion__curso')
    
    notas_dict = {}
    for n in notas_db:
        if n.valor is not None:
            if n.matricula_id not in notas_dict:
                notas_dict[n.matricula_id] = {}
            curso_id = n.evaluacion.asignacion.curso.id
            if curso_id not in notas_dict[n.matricula_id]:
                notas_dict[n.matricula_id][curso_id] = []
            notas_dict[n.matricula_id][curso_id].append(float(n.valor))

    datos_reporte = []
    for mat in matriculas:
        cursos_alumno = []
        notas_alumno = notas_dict.get(mat.id, {})
        todas_las_notas = [] # 💥 Para guardar todas las notas y sacar el promedio global
        
        for asig in asignaciones:
            notas_curso = notas_alumno.get(asig.curso.id, [])
            if notas_curso: 
                todas_las_notas.extend(notas_curso)
                cursos_alumno.append({
                    'nombre': asig.curso.nombre,
                    'notas': notas_curso,
                })

        if cursos_alumno:
            # 💥 Calculamos el promedio global del tema
            promedio_global = round(sum(todas_las_notas) / len(todas_las_notas), 2) if todas_las_notas else 0
            
            datos_reporte.append({
                'estudiante': f"{mat.estudiante.apellidos}, {mat.estudiante.nombres}",
                'cursos': cursos_alumno,
                'promedio_global': promedio_global
            })

    if request.GET.get('excel') == '1':
        return exportar_agenda_excel(aula, tema_seleccionado, datos_reporte)

    return render(request, 'personal/reporte_agenda.html', {
        'aula': aula,
        'temas_disponibles': temas_disponibles,
        'tema_seleccionado': tema_seleccionado,
        'datos_reporte': datos_reporte
    })

def exportar_agenda_excel(aula, tema, datos_reporte):
    """ Dibuja los tickets con 2 columnas y el Promedio General arriba """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets de Agenda"
    ws.views.sheetView[0].showGridLines = False

    font_titulo = Font(bold=True, size=12, name="Arial")
    font_bold = Font(bold=True, size=10, name="Arial")
    font_normal = Font(size=10, name="Arial")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    fila_actual = 2

    for alumno in datos_reporte:
        ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=3)
        ws.cell(row=fila_actual, column=2, value="REPORTE DE DESAFÍOS DIARIOS").font = font_titulo
        ws.cell(row=fila_actual, column=2).alignment = align_center
        fila_actual += 1

        ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=3)
        ws.cell(row=fila_actual, column=2, value=f"TEMA: {tema} | PROM. GLOBAL: {alumno['promedio_global']}").font = font_bold
        ws.cell(row=fila_actual, column=2).alignment = align_center
        fila_actual += 2

        ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=3)
        ws.cell(row=fila_actual, column=2, value=f"Estudiante: {alumno['estudiante']}").font = font_normal
        fila_actual += 1

        ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=3)
        ws.cell(row=fila_actual, column=2, value=f"Aula: {aula.grado} '{aula.seccion}'").font = font_normal
        fila_actual += 2

        cabeceras = ["CURSO", "NOTA"]
        for col_idx, cab in enumerate(cabeceras, start=2):
            c = ws.cell(row=fila_actual, column=col_idx, value=cab)
            c.font = font_bold
            c.border = borde
            c.alignment = align_center

        fila_actual += 1
        for curso in alumno['cursos']:
            notas_str = " - ".join([str(int(n)) if n.is_integer() else str(n) for n in curso['notas']])
            c1 = ws.cell(row=fila_actual, column=2, value=curso['nombre'])
            c2 = ws.cell(row=fila_actual, column=3, value=notas_str)

            for c in (c1, c2): c.border = borde
            c1.alignment = align_left
            c2.alignment = align_center
            fila_actual += 1

        fila_actual += 2
        ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=3)
        ws.cell(row=fila_actual, column=2, value="_______________________________").alignment = align_center
        fila_actual += 1
        ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=3)
        ws.cell(row=fila_actual, column=2, value="Firma del Apoderado").alignment = align_center

        fila_actual += 3
        ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=4)
        ws.cell(row=fila_actual, column=1, value="-------------------------------- ✂️ Línea de corte --------------------------------").alignment = align_center
        fila_actual += 3

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Tickets_{aula.grado}_{aula.seccion}_Tema_{tema[:15]}.xlsx"'
    wb.save(response)
    return response

@login_required
def sincronizar_google_sheets(request, aula_id):
    print(f"========== INICIANDO SINCRONIZACIÓN PARA EL AULA {aula_id} ==========")
    aula = get_object_or_404(Aula, id=aula_id)

    if not aula.google_sheet_id:
        messages.error(request, "Esta aula no tiene un ID de Google Sheets configurado.")
        return redirect('academico:mi_aula') 

    try:
        cliente = gspread.service_account(filename='credenciales_google.json')
        libro = cliente.open_by_key(aula.google_sheet_id)

        try:
            hoja = libro.worksheet('Datos_Django')
        except gspread.exceptions.WorksheetNotFound:
            hoja = libro.add_worksheet(title='Datos_Django', rows="100", cols="100") 

        asignaciones = AsignacionAcademica.objects.filter(aula=aula).select_related('curso')
        matriculas = Matricula.objects.filter(aula=aula, estudiante__estado='Activo').select_related('estudiante').order_by('estudiante__apellidos')
        
        # 💥 1. DEFINIMOS LOS NOMBRES EXACTOS COMO ESTÁN EN LA BASE DE DATOS
        TIPOS_NOTAS = [
            'Libro - Revisión 1', 'Cuaderno - Revisión 1',
            'Libro - Revisión 2', 'Cuaderno - Revisión 2',
            'Desafío - Tema 1', 'Desafío - Tema 2', 'Desafío - Tema 3', 'Desafío - Tema 4',
            'Desafío - Tema 5', 'Desafío - Tema 6', 'Desafío - Tema 7', 'Desafío - Tema 8',
            'Control de Calidad', 'ISO Ingeniería',
            'Concurso de Aptitud Mensual 1', 'Concurso de Aptitud Mensual 2'
        ]
        bimestres = ['I', 'II', 'III', 'IV']

        # 💥 1. CABECERAS (Solo 67 columnas: Clave + Alumno + Curso + 64 notas)
        cabeceras = ['CLAVE_BUSQUEDA', 'APELLIDOS Y NOMBRES', 'CURSO']
        for b in bimestres:
            for tipo in TIPOS_NOTAS:
                cabeceras.append(f"{b} - {tipo}")
                
        datos_a_subir = [cabeceras]

        # 💥 2. DICCIONARIO EN MEMORIA
        
        # PRO-TIP: Filtramos también por el Periodo Lectivo Activo 
        # para no mezclar notas del 2025 con las del 2026.
        periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
        
        notas_db = Nota.objects.filter(
            matricula__aula=aula,
            matricula__periodo=periodo_actual # <-- Evita bugs a futuro
        ).select_related('evaluacion', 'evaluacion__asignacion')

        print("\n========== ESCÁNER DE EVALUACIONES EN BD ==========")
        diccionario_notas = {}
        for n in notas_db:
            m_id = n.matricula_id
            c_id = n.evaluacion.asignacion.curso_id
            b_val = n.evaluacion.bimestre 
            
            # Limpieza del texto
            # nombre_eval = n.evaluacion.nombre.strip().upper().replace('Í', 'I')
            # Tomamos el nombre exacto, tal cual lo guardó el AJAX
            nombre_eval = n.evaluacion.nombre.strip()
            
            # 💥 ESTE PRINT TE DIRÁ LA VERDAD ABSOLUTA
            print(f"Bimestre: '{b_val}' | Curso: '{n.evaluacion.asignacion.curso.nombre}' | Evaluación: '{nombre_eval}' | Nota: {n.valor}")

            if m_id not in diccionario_notas: diccionario_notas[m_id] = {}
            if c_id not in diccionario_notas[m_id]: diccionario_notas[m_id][c_id] = {}
            if b_val not in diccionario_notas[m_id][c_id]: diccionario_notas[m_id][c_id][b_val] = {}

            diccionario_notas[m_id][c_id][b_val][nombre_eval] = float(n.valor) if n.valor is not None else ""
        print("===================================================\n")
        
        # 💥 3. ARMAR FILAS VERTICALMENTE POR CURSO
        for asig in asignaciones:
            curso_nombre = asig.curso.nombre
            
            for mat in matriculas:
                alumno_nombre = f"{mat.estudiante.apellidos}, {mat.estudiante.nombres}"
                
                # Fila Base
                fila = [
                    f"{alumno_nombre} - {curso_nombre}", # Col A: La Clave Secreta para el BUSCARV
                    alumno_nombre,                       # Col B: Nombre normal
                    curso_nombre                         # Col C: Curso
                ]
                
                # Rellenar las 64 notas horizontalmente para este alumno en este curso
                for b in bimestres:
                    for tipo in TIPOS_NOTAS:
                        nota = diccionario_notas.get(mat.id, {}).get(asig.curso.id, {}).get(b, {}).get(tipo, "")
                        fila.append(nota)
                
                datos_a_subir.append(fila)

        # 💥 4. SUBIR A DRIVE
        hoja.clear() 
        hoja.update(values=datos_a_subir, range_name='A1') 

        messages.success(request, f"¡Sincronización exitosa con Drive para el Aula {aula.grado} '{aula.seccion}'!")

    except Exception as e:
        print(f"====== ERROR AL SINCRONIZAR: {str(e)} ======")
        messages.error(request, f"Error al sincronizar con Google Drive. Hubo un problema de conexión.")

    return redirect('academico:mi_aula')

@login_required
def mi_perfil(request):
    # Intentamos obtener el perfil del personal asociado al usuario autenticado
    try:
        perfil = request.user.perfil_personal
    except Personal.DoesNotExist:
        messages.error(request, "No tienes un perfil de personal asignado a esta cuenta.")
        return redirect('index')

    # Inicializamos los formularios en estado base (GET)
    perfil_form = EditarPerfilForm(instance=perfil)
    password_form = PasswordChangeForm(user=request.user)

    if request.method == 'POST':
        # CASO A: El usuario decidió actualizar sus datos de contacto
        if 'btn_actualizar_datos' in request.POST:
            perfil_form = EditarPerfilForm(request.POST, instance=perfil)
            if perfil_form.is_valid():
                perfil_form.save()
                messages.success(request, '¡Datos de contacto actualizados correctamente!')
                return redirect('personal:mi_perfil')
            else:
                messages.error(request, 'Hubo un error al actualizar tus datos. Verifica el formulario.')

        # CASO B: El usuario decidió cambiar su contraseña
        elif 'btn_cambiar_password' in request.POST:
            password_form = PasswordChangeForm(user=request.user, data=request.POST)
            if password_form.is_valid():
                user = password_form.save()
                # 💥 Truco maestro: Esto evita que Django le cierre la sesión al usuario tras cambiar la clave
                update_session_auth_hash(request, user)
                messages.success(request, '¡Tu contraseña ha sido cambiada con éxito!')
                return redirect('personal:mi_perfil')
            else:
                messages.error(request, 'Por favor, corrige los errores en el formulario de credenciales.')

    context = {
        'perfil': perfil,
        'perfil_form': perfil_form,
        'password_form': password_form,
        'segment': 'perfil', # Mantiene iluminada la opción en tu sidebar
    }
    return render(request, 'personal/mi_perfil.html', context)

@login_required
def auditoria_academica_admin(request):
    """ Panel Panóptico para monitorear el cierre de notas de los docentes """
    personal_actual = obtener_personal_logueado(request)
    if personal_actual.cargo not in ['DIR', 'COO']:
        return redirect('home') # Seguridad: Solo directivos

    # 1. Primero obtenemos el periodo lectivo que está activo en el colegio
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()
    
    # 💥 LA SOLUCIÓN: Si hay un periodo activo, extraemos su bimestre actual.
    # Si por alguna razón no hay ninguno activo, usamos 'I' como salvavidas.
    bimestre_predeterminado = periodo_actual.bimestre_actual if periodo_actual else 'I'
    
    # 2. Capturamos el parámetro de la URL. Si viene vacío, adoptará el del sistema (ej: 'II')
    bimestre_seleccionado = request.GET.get('bimestre', bimestre_predeterminado)
    periodo_actual = PeriodoLectivo.objects.filter(activo=True).first()

    # 1. Traemos toda la carga académica del año
    asignaciones = AsignacionAcademica.objects.filter(
        periodo=periodo_actual
    ).select_related('aula', 'curso', 'personal').order_by('aula__nivel', 'aula__grado', 'curso__nombre')

    # 2. Traemos los candados (Cierres) de este bimestre
    cierres = CierreRegistroBimestral.objects.filter(
        bimestre=bimestre_seleccionado, 
        asignacion__periodo=periodo_actual
    )
    dict_cierres = {c.asignacion_id: c for c in cierres}

    # 3. Traemos las asignaciones que ya tienen al menos 1 evaluación creada en este bimestre (Para saber quién está trabajando)
    asignaciones_con_evaluaciones = set(
        Evaluacion.objects.filter(
            bimestre=bimestre_seleccionado, 
            asignacion__periodo=periodo_actual
        ).values_list('asignacion_id', flat=True)
    )

    # 4. Construimos la Matriz Inteligente
    lista_auditoria = []
    for asig in asignaciones:
        cierre = dict_cierres.get(asig.id)
        
        if cierre and cierre.cerrado:
            estado = 'CERRADO / ENTREGADO'
            color = 'success'
            icono = 'check_circle'
        elif asig.id in asignaciones_con_evaluaciones:
            estado = 'EN PROGRESO'
            color = 'warning'
            icono = 'hourglass_bottom'
        else:
            estado = 'SIN INICIAR'
            color = 'danger'
            icono = 'cancel'

        lista_auditoria.append({
            'asignacion': asig,
            'estado': estado,
            'color': color,
            'icono': icono,
        })

    # Extraemos la lista de bimestres desde el modelo PeriodoLectivo
    opciones_bimestres = PeriodoLectivo.BIMESTRES
    
    # Lo que HAN cumplido
    solicitudes = SolicitudImpresion.objects.filter(asignacion__periodo=periodo_actual).prefetch_related('archivos').order_by('-fecha_subida')
    
    # 💥 Obtenemos los simulacros completados de forma rápida
    entregas_simulacro = set(EntregaSimulacro.objects.filter(finalizado=True).values_list('curso_id', 'docente_id'))
    
    matriz_cumplimiento = []
    for asig in asignaciones:
        sols_asig = [s for s in solicitudes if s.asignacion_id == asig.id]
        
        # Empaquetamos qué temas y qué exámenes ya subió este profesor
        temas_dict = {}
        for s in sols_asig:
            temas_dict[s.tema] = True # Registramos el Tema (TEMA_1, TEMA_2, etc.)
            
            # Revisamos dentro de los archivos por si hay un examen
            tipos_archivos = [a.tipo for a in s.archivos.all()]
            if 'CALIDAD' in tipos_archivos:
                temas_dict['CALIDAD'] = True
            if 'ISO' in tipos_archivos:
                temas_dict['ISO'] = True
                
        tiene_simulacro = (asig.curso_id, asig.personal_id) in entregas_simulacro
        
        # 💥 SEMÁFORO GLOBAL DE COLORES
        if sols_asig and tiene_simulacro:
            estado_color = 'success' # Verde (Todo al día)
            icono = 'check_circle'
        elif sols_asig:
            estado_color = 'warning' # Naranja (En progreso)
            icono = 'schedule'
        else:
            estado_color = 'danger'  # Rojo (Falta todo)
            icono = 'cancel'
            
        matriz_cumplimiento.append({
            'asignacion_id': asig.id,
            'docente': f"{asig.personal.apellidos}, {asig.personal.nombres}",
            'aula_texto': f"{asig.aula.nivel} - {asig.aula.grado} {asig.aula.seccion}",
            'curso': asig.curso.nombre,
            'estado_color': estado_color,
            'icono': icono,
            'total_envios': len(sols_asig),
            'temas_json': json.dumps(temas_dict), # Magia pura para el Frontend
            'tiene_simulacro': tiene_simulacro,
        })

    return render(request, 'personal/auditoria_academica.html', {
        'lista_auditoria': lista_auditoria,
        'bimestre_seleccionado': bimestre_seleccionado,
        'opciones_bimestres': opciones_bimestres,
        'matriz_cumplimiento': matriz_cumplimiento,
    })